from flask import Flask, render_template, request, jsonify, send_file, send_from_directory, redirect, url_for, flash
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from flask_bcrypt import Bcrypt
import threading
import uuid
import json
import os
import logging
import sys
import pandas as pd
import numpy as np
from datetime import datetime, timezone, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO
from tms_database import get_tms_orders_by_day_fedex, get_tms_orders_by_day_supplier, get_tms_supplier_product_names, get_tms_supplier_records_for_verification
from fedex_db import (
    init_fedex_db,
    upsert_fedex_shipment,
    bulk_upsert_fedex_shipments,
    get_dashboard_customers,
    get_dashboard_orders_by_weight,
    get_dashboard_orders_by_zone,
    get_dashboard_surcharge_stats,
    get_dashboard_adjustment_stats
)
from user_db import (
    init_user_db,
    get_user_by_id,
    get_user_by_username,
    get_password_hash,
    create_user,
    update_user,
    update_password,
    update_last_login,
    log_login_attempt,
    get_all_users,
    delete_user,
    create_default_admin,
    AVAILABLE_PERMISSIONS,
    DEFAULT_ROLE_PERMISSIONS,
    PERMISSION_GROUPS
)

# Define FEDEX_DB_PATH at the top level so it is available everywhere
FEDEX_DB_PATH = os.path.join(os.path.dirname(__file__), 'fedex_data.db')

EXPORTS_DIR = os.path.join(os.path.dirname(__file__), 'exports')
EXPORTS_METADATA_FILE = os.path.join(EXPORTS_DIR, 'exports_metadata.json')
os.makedirs(EXPORTS_DIR, exist_ok=True)
export_jobs = {}  # job_id -> {'status': 'pending'/'done'/'error', 'file': filename, 'error': str}

# Store last analysis results per session for efficient export (avoids large POST payloads)
analysis_results_cache = {}  # session_id -> {'tms_comparison': [...], 'unfound_in_tms': [...], 'timestamp': ...}

# Frontend lives one level up in the repo root
_REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
app = Flask(
    __name__,
    static_folder=os.path.join(_REPO_ROOT, 'frontend', 'static'),
    template_folder=os.path.join(_REPO_ROOT, 'frontend', 'templates'),
)

# Flask-Login and Bcrypt setup
app.secret_key = os.environ.get('SECRET_KEY', 'logistar-platform-secret-key-change-in-production')
bcrypt = Bcrypt(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Please log in to access this page.'
login_manager.login_message_category = 'info'

# Session configuration
app.config['REMEMBER_COOKIE_DURATION'] = timedelta(days=7)
app.config['SESSION_PROTECTION'] = 'strong'

# Register WMS Monitor and Turnover Analytics blueprints
from wms_blueprint import wms_bp
from turnover_blueprint import turnover_bp
app.register_blueprint(wms_bp)
app.register_blueprint(turnover_bp)

# Background scheduler (WMS monitor + TMS daily export)
from scheduler import start_scheduler, stop_scheduler, get_scheduler_status, trigger_job_now


@login_manager.user_loader
def load_user(user_id):
    """Load user by ID for Flask-Login"""
    return get_user_by_id(int(user_id))


def load_exports_metadata():
    """Load export metadata from JSON file"""
    if os.path.exists(EXPORTS_METADATA_FILE):
        try:
            with open(EXPORTS_METADATA_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            logging.error(f"Error loading exports metadata: {e}")
            return []
    return []

def save_exports_metadata(metadata):
    """Save export metadata to JSON file"""
    try:
        with open(EXPORTS_METADATA_FILE, 'w', encoding='utf-8') as f:
            json.dump(metadata, f, indent=2, ensure_ascii=False)
    except Exception as e:
        logging.error(f"Error saving exports metadata: {e}")

def add_export_metadata(original_filename, export_type, file_path, file_size, row_count=0, exported_by_id=None, exported_by_username=None):
    """Add metadata for a new export file"""
    metadata = load_exports_metadata()
    metadata.append({
        'id': str(uuid.uuid4()),
        'original_filename': original_filename,
        'export_type': export_type,
        'file_path': file_path,
        'file_size': file_size,
        'row_count': row_count,
        'exported_by_id': exported_by_id,
        'exported_by_username': exported_by_username or 'unknown',
        'timestamp': datetime.now(timezone.utc).isoformat()
    })
    save_exports_metadata(metadata)
    return metadata[-1]


def _create_not_found_sheet(ws, rows):
    """Helper: Create sheet for records not found in TMS.
    
    Args:
        ws: Worksheet object
        rows: List of unfound records
    """
    headers = [
        'Tracking ID', 'Master Tracking', 'Zone', 'Weight (lbs)', 'FedEx Charge',
        'Ship Date', 'Dimensions', 'Service Type'
    ]
    ws.append(headers)
    
    # Style header row
    from openpyxl.styles import PatternFill, Font, Alignment
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add data rows
    for record in rows:
        ws.append([
            record.get('tracking_id', ''),
            record.get('TDMasterTrackingID', '').lstrip("'"),
            record.get('zone', ''),
            record.get('actual_weight', 0),
            record.get('charge', 0),
            record.get('ship_date', ''),
            record.get('dimensions', ''),
            record.get('service_type', '')
        ])


def _identify_discount_columns(rows, original_column_names):
    """Helper: Identify discount columns that should be excluded from export.
    
    Args:
        rows: List of customer records
        original_column_names: List of column names in original order
        
    Returns:
        Set of column names to exclude
    """
    excluded_columns = set()
    charge_cols_to_check = [col for col in original_column_names if 'Tracking ID Charge Amount' in col]
    negative_charge_cols = set()
    
    # Single pass through rows to check all charge columns at once
    for record in rows:
        orig_row = record.get('original_row', {})
        for col_name in charge_cols_to_check:
            if col_name in negative_charge_cols:
                continue  # Already identified as negative
            amount_val = orig_row.get(col_name, '')
            try:
                if amount_val and float(amount_val) < 0:
                    negative_charge_cols.add(col_name)
            except (ValueError, TypeError):
                pass
        
        # Early exit if all charge columns checked
        if len(negative_charge_cols) == len(charge_cols_to_check):
            break
    
    # Add discount columns and their descriptions to excluded set
    for col_name in negative_charge_cols:
        excluded_columns.add(col_name)
        desc_col = col_name.replace('Amount', 'Description')
        excluded_columns.add(desc_col)
    
    return excluded_columns


def _build_filtered_columns(original_column_names, excluded_columns):
    """Helper: Build list of columns starting from Service Type, excluding discounts.
    
    Args:
        original_column_names: List of all column names in order
        excluded_columns: Set of columns to exclude
        
    Returns:
        List of filtered column names
    """
    all_original_columns = []
    service_type_found = False
    
    for col_name in original_column_names:
        if service_type_found:
            if col_name not in excluded_columns:
                all_original_columns.append(col_name)
        elif col_name == 'Service Type':
            service_type_found = True
            all_original_columns.append(col_name)
    
    return all_original_columns


def background_export_customer_adjustments(job_id, tms_comparison, unfound_in_tms, original_filename=None, customers_dict=None, exported_by_id=None, exported_by_username=None):
    """Background task to export customer adjustment reports to Excel.
    
    Args:
        job_id: Unique identifier for this export job
        tms_comparison: List of comparison records with customer data
        unfound_in_tms: List of records not found in TMS
        original_filename: Name of original uploaded invoice file
        customers_dict: Dictionary mapping customer names to their pricing configuration
        exported_by_id: User ID who initiated the export
        exported_by_username: Username who initiated the export
    """
    try:
        logging.info(f"[EXPORT {job_id}] Starting export for {len(tms_comparison)} records")
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from datetime import datetime
        
        # Initialize customers_dict if not provided
        if customers_dict is None:
            customers_dict = {}
        
        wb = Workbook()
        wb.remove(wb.active)
        customers_data = {}
        for row in tms_comparison:
            customer = row.get('customer', 'Unknown')
            if customer not in customers_data:
                customers_data[customer] = []
            customers_data[customer].append(row)
        if unfound_in_tms:
            customers_data['Not Found in TMS'] = unfound_in_tms
        for customer_name, rows in customers_data.items():
            ws = wb.create_sheet(title=customer_name[:31])
            if customer_name == 'Not Found in TMS':
                _create_not_found_sheet(ws, rows)
            else:
                # Use first row to get column structure
                sample_row = rows[0] if rows else {}
                original_row = sample_row.get('original_row', {})
                original_column_names = sample_row.get('original_column_order', [])
                if not original_column_names:
                    original_column_names = list(original_row.keys()) if original_row else []
                
                # Identify discount columns to exclude (single-pass optimization)
                excluded_columns = _identify_discount_columns(rows, original_column_names)
                
                # Build filtered column list starting from Service Type
                all_original_columns = _build_filtered_columns(original_column_names, excluded_columns)
                
                # Build headers: our calculated columns + filtered original invoice columns
                headers = [
                    'Tracking ID',
                    'Base Rate\n基础费用',
                    'Surcharge\n附加费',
                    'Prepaid Charge\n系统已收',
                    'Total Amount\n应收费用',
                    'Adjustment\n差异调整',
                ] + all_original_columns
                
                ws.append(headers)
                header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
                header_font = Font(name='Arial', size=10, bold=True)
                for idx, cell in enumerate(ws[1], 1):
                    if idx == 6:
                        cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                ws.row_dimensions[1].height = 30
                for record in rows:
                    we_should_charge = record.get('expected_customer_charge', None)
                    tms_charged = record.get('tms_charged', 0)
                    adjustment = (we_should_charge or 0) - tms_charged
                    base_rate = record.get('base_rate', '')
                    customer_name = record.get('customer', '')
                    
                    # Get customer markup information for applying to charge amounts
                    customer_markup_info = None
                    if customer_name in customers_dict:
                        customer_markup_info = customers_dict[customer_name]
                    
                    # Build row with calculated values + original invoice columns
                    row_data = [
                        record.get('tracking_id', ''),
                        base_rate,
                        record.get('surcharge', ''),
                        tms_charged,
                        we_should_charge if we_should_charge is not None and str(we_should_charge) != '' else '',
                        adjustment,
                    ]
                    
                    # Add original invoice columns (filtered) with markup applied to charge amounts
                    original_row = record.get('original_row', {})
                    for col_name in all_original_columns:
                        cell_value = original_row.get(col_name, '')
                        
                        # Apply customer markup to Tracking ID Charge Amount columns
                        if 'Tracking ID Charge Amount' in col_name and customer_markup_info and cell_value:
                            try:
                                charge_amount = float(cell_value)
                                # Get corresponding description to determine markup type
                                desc_col = col_name.replace('Amount', 'Description')
                                desc = original_row.get(desc_col, '').lower()
                                
                                # Determine markup percentage based on surcharge type
                                if 'fuel' in desc:
                                    markup_pct = customer_markup_info.get('fuel_markup', 0)
                                elif any(kw in desc for kw in ['zone', 'extended', 'delivery area']):
                                    markup_pct = customer_markup_info.get('zone_markup', 0)
                                elif any(kw in desc for kw in ['demand', 'peak']):
                                    markup_pct = customer_markup_info.get('demand_markup', 0)
                                else:
                                    markup_pct = customer_markup_info.get('fixed_markup', 0)
                                
                                # Apply markup and round to 2 decimal places
                                cell_value = round(charge_amount * (1 + markup_pct / 100), 2)
                            except (ValueError, TypeError):
                                pass  # Keep original value if conversion fails
                        
                        row_data.append(cell_value)
                    
                    ws.append(row_data)
                    current_row = ws.max_row
                    for idx, cell in enumerate(ws[current_row], 1):
                        cell.font = Font(name='Arial', size=10)
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                        if idx in [2, 4, 5, 6]:
                            cell.number_format = '$#,##0.00'
            
            # Set column widths - first 6 columns are our calculated columns
            base_column_widths = {
                'A': 18,  # Tracking ID
                'B': 15,  # Base Rate
                'C': 30,  # Surcharge
                'D': 15,  # Prepaid Charge
                'E': 18,  # Total Amount
                'F': 12,  # Adjustment
            }
            for col_letter, width in base_column_widths.items():
                ws.column_dimensions[col_letter].width = width
            
            # Set default width for remaining columns (original invoice columns)
            if customer_name != 'Not Found in TMS':
                from openpyxl.utils import get_column_letter
                for col_idx in range(7, len(headers) + 1):
                    col_letter = get_column_letter(col_idx)
                    ws.column_dimensions[col_letter].width = 12
        
        # Generate filename with timestamp and original filename
        timestamp = datetime.now().strftime('%Y-%m-%d_%H%M%S')
        base_name = original_filename.replace('.xlsx', '').replace('.xls', '').replace('.csv', '') if original_filename else 'invoice'
        filename = os.path.join(EXPORTS_DIR, f'{timestamp}_{base_name}_adjustment.xlsx')
        wb.save(filename)
        
        # Add metadata
        file_size = os.path.getsize(filename)
        total_rows = len(tms_comparison) + (len(unfound_in_tms) if unfound_in_tms else 0)
        add_export_metadata(
            original_filename=original_filename or 'unknown',
            export_type='adjustment',
            file_path=filename,
            file_size=file_size,
            row_count=total_rows,
            exported_by_id=exported_by_id,
            exported_by_username=exported_by_username
        )
        
        export_jobs[job_id]['status'] = 'done'
        export_jobs[job_id]['file'] = filename
        export_jobs[job_id]['filename'] = os.path.basename(filename)
        logging.info(f"[EXPORT {job_id}] Completed: {filename} ({total_rows} rows)")
    except Exception as e:
        logging.error(f"[EXPORT {job_id}] Failed: {e}")
        export_jobs[job_id]['status'] = 'error'
        export_jobs[job_id]['error'] = str(e)


# Configure logging to show TMS scraper progress
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s %(levelname)s: %(message)s',
    handlers=[
        logging.StreamHandler(sys.stdout)
    ]
)

DATA_DIR = os.path.join(os.path.dirname(__file__), 'data')
DATA_FILE = os.path.join(DATA_DIR, 'rates.json')
CUSTOMER_PRICING_FILE = os.path.join(DATA_DIR, 'customer_pricing.json')


def ensure_data_dir():
    if not os.path.exists(DATA_DIR):
        os.makedirs(DATA_DIR)


def load_data():
    """Load all rate versions and return the full versioned structure"""
    ensure_data_dir()
    if not os.path.exists(DATA_FILE):
        # default empty versioned structure
        zones = [f'Zone {z}' for z in range(2, 9)]
        lbs = [i for i in range(1, 151)]
        empty_version = {
            'base_table': {
                'zones': zones,
                'lbs': lbs,
                'rates': [["" for _ in zones] for _ in lbs]
            },
            'fixed_surcharges': [],
            'zone_based_surcharges': [],
            'demand_surcharges': [],
            'fuel_surcharge': {'date_ranges': []}
        }
        data = {
            'rate_versions': [
                {
                    'version_id': 'v1_default',
                    'version_name': 'Default Rates',
                    'effective_start': None,
                    'effective_end': None,
                    **empty_version
                }
            ],
            'active_version_id': 'v1_default'
        }
        save_data(data)
        return data
    
    with open(DATA_FILE, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    # Check if this is old non-versioned format
    if 'rate_versions' not in data:
        # Migrate old format to versioned
        logging.info("Migrating old rate format to versioned structure...")
        old_data = data
        data = {
            'rate_versions': [
                {
                    'version_id': 'v1_legacy',
                    'version_name': 'Legacy Rates',
                    'effective_start': None,
                    'effective_end': None,
                    'base_table': old_data.get('base_table', {}),
                    'fixed_surcharges': old_data.get('fixed_surcharges', []),
                    'zone_based_surcharges': old_data.get('zone_based_surcharges', []),
                    'demand_surcharges': old_data.get('demand_surcharges', []),
                    'fuel_surcharge': old_data.get('fuel_surcharge', {'date_ranges': []})
                }
            ],
            'active_version_id': 'v1_legacy'
        }
        save_data(data)
    
    # Normalize all versions
    fixed_zones = [f'Zone {z}' for z in range(2, 9)]
    fixed_lbs = [i for i in range(1, 151)]
    
    for version in data.get('rate_versions', []):
        bt = version.get('base_table', {})
        old_rates = bt.get('rates', [])
        
        # Normalize rates matrix
        normalized = []
        for r_idx in range(len(fixed_lbs)):
            row = []
            for c_idx in range(len(fixed_zones)):
                val = ''
                if r_idx < len(old_rates) and c_idx < len(old_rates[r_idx]):
                    val = old_rates[r_idx][c_idx]
                row.append(val)
            normalized.append(row)
        
        version['base_table'] = {
            'zones': fixed_zones,
            'lbs': fixed_lbs,
            'rates': normalized
        }
        
        # Ensure all keys exist
        if 'fixed_surcharges' not in version:
            version['fixed_surcharges'] = []
        if 'zone_based_surcharges' not in version:
            version['zone_based_surcharges'] = []
        if 'demand_surcharges' not in version:
            version['demand_surcharges'] = []
        if 'fuel_surcharge' not in version:
            version['fuel_surcharge'] = {'date_ranges': []}
    
    return data


def get_version_by_date(shipment_date):
    """Get the appropriate rate version based on shipment date"""
    data = load_data()
    
    # If shipment_date is None, use active version
    if not shipment_date:
        active_id = data.get('active_version_id')
        for version in data.get('rate_versions', []):
            if version.get('version_id') == active_id:
                return version
        # Fallback to first version
        return data.get('rate_versions', [{}])[0] if data.get('rate_versions') else {}
    
    # Parse shipment date if it's a string
    if isinstance(shipment_date, str):
        try:
            shipment_date = datetime.strptime(shipment_date, '%Y-%m-%d').date()
        except:
            shipment_date = None
    elif isinstance(shipment_date, datetime):
        shipment_date = shipment_date.date()
    
    if not shipment_date:
        # Fallback to active version
        active_id = data.get('active_version_id')
        for version in data.get('rate_versions', []):
            if version.get('version_id') == active_id:
                return version
        return data.get('rate_versions', [{}])[0] if data.get('rate_versions') else {}
    
    # Find version where shipment_date falls within effective range
    for version in data.get('rate_versions', []):
        start_str = version.get('effective_start')
        end_str = version.get('effective_end')
        
        # Parse dates
        start_date = None
        end_date = None
        
        if start_str:
            try:
                start_date = datetime.strptime(start_str, '%Y-%m-%d').date()
            except:
                pass
        
        if end_str:
            try:
                end_date = datetime.strptime(end_str, '%Y-%m-%d').date()
            except:
                pass
        
        # Check if date falls in range
        if start_date and shipment_date < start_date:
            continue
        if end_date and shipment_date > end_date:
            continue
        
        return version
    
    # No matching version found, use active version as fallback
    active_id = data.get('active_version_id')
    for version in data.get('rate_versions', []):
        if version.get('version_id') == active_id:
            return version
    
    return data.get('rate_versions', [{}])[0] if data.get('rate_versions') else {}


def get_version_by_id(version_id):
    """Get a specific rate version by ID"""
    data = load_data()
    for version in data.get('rate_versions', []):
        if version.get('version_id') == version_id:
            return version
    return None


def save_data(data):
    ensure_data_dir()
    with open(DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


# ===============================
# Authentication Routes
# ===============================

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Login page and authentication"""
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    
    error = None
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        remember = request.form.get('remember') == '1'
        
        user = get_user_by_username(username)
        password_hash = get_password_hash(username)
        
        if user and password_hash and bcrypt.check_password_hash(password_hash, password):
            if not user.is_active:
                error = 'Your account has been disabled. Please contact an administrator.'
                log_login_attempt(user.id, request.remote_addr, request.user_agent.string, success=False)
            else:
                login_user(user, remember=remember)
                update_last_login(user.id)
                log_login_attempt(user.id, request.remote_addr, request.user_agent.string, success=True)
                
                next_page = request.args.get('next')
                if next_page:
                    return redirect(next_page)
                return redirect(url_for('index'))
        else:
            error = 'Invalid username or password.'
            if user:
                log_login_attempt(user.id, request.remote_addr, request.user_agent.string, success=False)
    
    return render_template('login.html', error=error)


@app.route('/logout')
@login_required
def logout():
    """Logout user"""
    logout_user()
    return redirect(url_for('login'))


@app.route('/favicon.ico')
def favicon():
    """Serve favicon"""
    return send_from_directory(os.path.join(app.root_path, 'static'),
                               'logistar_logo.png', mimetype='image/png')


@app.route('/')
@login_required
def index():
    """Main application page - requires login"""
    return render_template('index.html', 
                           user=current_user,
                           permissions=current_user.get_permissions())


# ===============================
# User Management API Routes
# ===============================

@app.route('/api/users', methods=['GET'])
@login_required
def api_get_users():
    """Get all users (admin only)"""
    if not current_user.has_permission('user_management'):
        return jsonify({'success': False, 'error': 'Permission denied'}), 403
    
    users = get_all_users()
    return jsonify({
        'success': True,
        'users': [u.to_dict() for u in users],
        'available_permissions': AVAILABLE_PERMISSIONS,
        'default_role_permissions': DEFAULT_ROLE_PERMISSIONS,
        'permission_groups': PERMISSION_GROUPS
    })


@app.route('/api/users', methods=['POST'])
@login_required
def api_create_user():
    """Create a new user (admin only)"""
    if not current_user.has_permission('user_management'):
        return jsonify({'success': False, 'error': 'Permission denied'}), 403
    
    data = request.json
    username = data.get('username', '').strip()
    password = data.get('password', '')
    email = data.get('email', '').strip()
    display_name = data.get('display_name', '').strip()
    role = data.get('role', 'viewer')
    permissions = data.get('permissions', DEFAULT_ROLE_PERMISSIONS.get(role, []))
    
    # Sub-admin can only create viewers
    if current_user.role == 'sub-admin' and role != 'viewer':
        return jsonify({'success': False, 'error': 'Sub-admin can only create Viewer users'}), 403
    
    if not username or not password:
        return jsonify({'success': False, 'error': 'Username and password are required'}), 400
    
    if len(password) < 6:
        return jsonify({'success': False, 'error': 'Password must be at least 6 characters'}), 400
    
    password_hash = bcrypt.generate_password_hash(password).decode('utf-8')
    user_id = create_user(
        username=username,
        password_hash=password_hash,
        email=email or None,
        display_name=display_name or username,
        role=role,
        permissions=permissions,
        created_by=current_user.id
    )
    
    if user_id:
        return jsonify({'success': True, 'user_id': user_id})
    else:
        return jsonify({'success': False, 'error': 'Username already exists'}), 400


@app.route('/api/users/<int:user_id>', methods=['PUT'])
@login_required
def api_update_user(user_id):
    """Update a user (admin only)"""
    if not current_user.has_permission('user_management'):
        return jsonify({'success': False, 'error': 'Permission denied'}), 403
    
    data = request.json
    
    # Sub-admin can only edit viewers
    if current_user.role == 'sub-admin':
        target_user = get_user_by_id(user_id)
        if target_user and target_user.role != 'viewer':
            return jsonify({'success': False, 'error': 'Sub-admin can only edit Viewer users'}), 403
        # Ensure sub-admin doesn't change role from viewer
        if data.get('role') and data.get('role') != 'viewer':
            return jsonify({'success': False, 'error': 'Sub-admin can only manage Viewer users'}), 403
    
    # Don't allow deactivating yourself
    if user_id == current_user.id and data.get('is_active') == False:
        return jsonify({'success': False, 'error': 'Cannot deactivate your own account'}), 400
    
    update_user(
        user_id=user_id,
        email=data.get('email'),
        display_name=data.get('display_name'),
        role=data.get('role'),
        permissions=data.get('permissions'),
        is_active=data.get('is_active')
    )
    
    return jsonify({'success': True})


@app.route('/api/users/<int:user_id>/password', methods=['PUT'])
@login_required
def api_update_user_password(user_id):
    """Update user password (admin or self)"""
    # Allow users to change their own password, or admin to change anyone's
    if user_id != current_user.id and not current_user.has_permission('user_management'):
        return jsonify({'success': False, 'error': 'Permission denied'}), 403
    
    data = request.json
    new_password = data.get('password', '')
    
    if len(new_password) < 6:
        return jsonify({'success': False, 'error': 'Password must be at least 6 characters'}), 400
    
    password_hash = bcrypt.generate_password_hash(new_password).decode('utf-8')
    update_password(user_id, password_hash)
    
    return jsonify({'success': True})


@app.route('/api/users/<int:user_id>', methods=['DELETE'])
@login_required
def api_delete_user(user_id):
    """Delete (deactivate) a user (admin only)"""
    if not current_user.has_permission('user_management'):
        return jsonify({'success': False, 'error': 'Permission denied'}), 403
    
    if user_id == current_user.id:
        return jsonify({'success': False, 'error': 'Cannot delete your own account'}), 400
    
    # Sub-admin can only delete viewers
    if current_user.role == 'sub-admin':
        target_user = get_user_by_id(user_id)
        if target_user and target_user.role != 'viewer':
            return jsonify({'success': False, 'error': 'Sub-admin can only delete Viewer users'}), 403
    
    delete_user(user_id)
    return jsonify({'success': True})


@app.route('/api/current-user', methods=['GET'])
@login_required
def api_get_current_user():
    """Get current logged-in user info"""
    return jsonify({
        'success': True,
        'user': current_user.to_dict()
    })


@app.route('/api/data', methods=['GET'])
@login_required
def api_get_data():
    """Get all rate versions"""
    data = load_data()
    return jsonify(data)


@app.route('/api/data/version/<version_id>', methods=['GET'])
def api_get_version(version_id):
    """Get a specific rate version"""
    version = get_version_by_id(version_id)
    if version:
        return jsonify(version)
    return jsonify({'error': 'Version not found'}), 404


@app.route('/api/data/active', methods=['GET'])
def api_get_active_version():
    """Get the currently active rate version for editing"""
    data = load_data()
    active_id = data.get('active_version_id')
    version = get_version_by_id(active_id)
    if version:
        return jsonify(version)
    return jsonify({'error': 'Active version not found'}), 404


@app.route('/api/data', methods=['POST'])
def api_save_data():
    """Save entire versioned rate structure"""
    data = request.get_json()
    if not data:
        return jsonify({'error': 'No JSON payload received'}), 400
    save_data(data)
    return jsonify({'status': 'ok'})


@app.route('/api/data/version/<version_id>', methods=['POST'])
def api_update_version(version_id):
    """Update a specific rate version"""
    updated_version = request.get_json()
    if not updated_version:
        return jsonify({'error': 'No JSON payload received'}), 400
    
    data = load_data()
    found = False
    
    for i, version in enumerate(data.get('rate_versions', [])):
        if version.get('version_id') == version_id:
            # Preserve version_id
            updated_version['version_id'] = version_id
            data['rate_versions'][i] = updated_version
            found = True
            break
    
    if not found:
        return jsonify({'error': 'Version not found'}), 404
    
    save_data(data)
    return jsonify({'status': 'ok'})


@app.route('/api/data/version', methods=['POST'])
def api_create_version():
    """Create a new rate version"""
    new_version = request.get_json()
    if not new_version:
        return jsonify({'error': 'No JSON payload received'}), 400
    
    data = load_data()
    
    # Ensure version has required fields
    if 'version_id' not in new_version:
        return jsonify({'error': 'version_id is required'}), 400
    
    # Check if version_id already exists
    for version in data.get('rate_versions', []):
        if version.get('version_id') == new_version['version_id']:
            return jsonify({'error': 'Version ID already exists'}), 400
    
    data['rate_versions'].append(new_version)
    save_data(data)
    return jsonify({'status': 'ok', 'version_id': new_version['version_id']})


@app.route('/api/data/version/<version_id>', methods=['DELETE'])
def api_delete_version(version_id):
    """Delete a rate version"""
    data = load_data()
    
    # Don't allow deletion if it's the only version
    if len(data.get('rate_versions', [])) <= 1:
        return jsonify({'error': 'Cannot delete the only rate version'}), 400
    
    # Don't allow deletion of active version
    if data.get('active_version_id') == version_id:
        return jsonify({'error': 'Cannot delete the active version. Please set another version as active first.'}), 400
    
    data['rate_versions'] = [v for v in data.get('rate_versions', []) if v.get('version_id') != version_id]
    save_data(data)
    return jsonify({'status': 'ok'})


@app.route('/api/data/active/<version_id>', methods=['POST'])
def api_set_active_version(version_id):
    """Set the active rate version"""
    data = load_data()
    
    # Check if version exists
    version_exists = any(v.get('version_id') == version_id for v in data.get('rate_versions', []))
    if not version_exists:
        return jsonify({'error': 'Version not found'}), 404
    
    data['active_version_id'] = version_id
    save_data(data)
    return jsonify({'status': 'ok'})


def load_customer_pricing():
    """Load customer pricing configuration"""
    ensure_data_dir()
    if not os.path.exists(CUSTOMER_PRICING_FILE):
        return {'customers': []}
    with open(CUSTOMER_PRICING_FILE, 'r', encoding='utf-8') as f:
        return json.load(f)


def save_customer_pricing(data):
    """Save customer pricing configuration"""
    ensure_data_dir()
    with open(CUSTOMER_PRICING_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


# Warehouse Operations Fees
WAREHOUSE_OPS_FILE = os.path.join(DATA_DIR, 'warehouse_ops.json')

def load_warehouse_ops():
    """Load warehouse operation fees configuration"""
    ensure_data_dir()
    if not os.path.exists(WAREHOUSE_OPS_FILE):
        return {
            'shelving_fees': {
                '0-2': 0, '2-5': 0, '5-10': 0, '10-15': 0,
                '15-20': 0, '20-25': 0, '25-30': 0, 'over-30': 0
            },
            'outbound_fees': {
                '0-1': 0, '1-2': 0, '2-5': 0, '5-10': 0, '10-15': 0,
                '15-20': 0, '20-25': 0, '25-30': 0, 'over-30': 0
            }
        }
    with open(WAREHOUSE_OPS_FILE, 'r', encoding='utf-8') as f:
        return json.load(f)


def save_warehouse_ops(data):
    """Save warehouse operation fees configuration"""
    ensure_data_dir()
    with open(WAREHOUSE_OPS_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


@app.route('/api/warehouse_ops', methods=['GET'])
def api_get_warehouse_ops():
    """Get warehouse operation fees configuration"""
    data = load_warehouse_ops()
    return jsonify(data)


@app.route('/api/warehouse_ops', methods=['POST'])
def api_save_warehouse_ops():
    """Save warehouse operation fees configuration"""
    data = request.get_json()
    if not data:
        return jsonify({'error': 'No JSON payload received'}), 400
    save_warehouse_ops(data)
    return jsonify({'status': 'ok'})


@app.route('/api/customers', methods=['GET'])
def api_get_customers():
    """Get customer pricing configuration"""
    data = load_customer_pricing()
    return jsonify(data)


@app.route('/api/customers', methods=['POST'])
def api_save_customers():
    """Save customer pricing configuration"""
    data = request.get_json()
    if not data:
        return jsonify({'error': 'No JSON payload received'}), 400
    save_customer_pricing(data)
    return jsonify({'status': 'ok'})


@app.route('/api/verify', methods=['POST'])
def api_verify_invoice():
    """Verify invoice line items against configured rates and surcharges"""
    try:
        invoice_data = request.get_json()
        
        if not invoice_data:
            return jsonify({'error': 'No JSON payload received'}), 400
        
        if 'rows' not in invoice_data:
            # Try to handle direct array format
            if isinstance(invoice_data, list):
                invoice_data = {'rows': invoice_data}
            else:
                return jsonify({
                    'error': 'Invalid invoice data format. Expected {"rows": [...]} or array. Got: ' + str(type(invoice_data))
                }), 400
        
        rows = invoice_data['rows']
        if not rows or len(rows) == 0:
            return jsonify({'error': 'No rows found in invoice data'}), 400
        
        num_rows = len(rows)
        logging.info(f"Processing {num_rows} invoice rows")
        
        # Store original column order from first row to maintain FedEx invoice column sequence
        # This is crucial because JSON serialization can reorder dictionary keys
        original_column_order = list(rows[0].keys()) if rows else []
        
        config = load_data()
        results = []
        
        # Phase 1: Verify each invoice row against our rates
        for idx, row in enumerate(rows):
            try:
                verification = verify_row(row, config)
                # Attach original row data and column order for export pipeline
                verification['original_row'] = row
                verification['original_column_order'] = original_column_order
                results.append(verification)
            except Exception as e:
                logging.warning(f"Error verifying row {idx}: {str(e)}")
                results.append({
                    'tracking_id': row.get('Express or Ground Tracking ID', f'Row {idx}'),
                    'error': str(e),
                    'status': 'error',
                    'original_row': row,
                    'original_column_order': original_column_order
                })
        
        # Phase 2: Perform TMS lookup to get customer info and prepaid charges
        logging.info(f"Phase 2: TMS lookup for {len([r for r in results if r.get('tracking_id')])} tracking numbers")
        tracking_numbers = []
        for r in results:
            if r.get('tracking_id'):
                tracking_num = str(r['tracking_id']).lstrip("'")  # Remove Excel text format apostrophe
                tracking_numbers.append(tracking_num)
        
        tms_data = {}
        
        try:
            # Try to query from database first (fast)
            from tms_database import get_records_by_tracking_numbers
            import os
            
            db_path = os.path.join(os.path.dirname(__file__), 'tms_data.db')
            
            if os.path.exists(db_path):
                tms_results = get_records_by_tracking_numbers(tracking_numbers)
                
                if tms_results:
                    # Create lookup dictionary by tracking number
                    for tms_row in tms_results:
                        tracking_num = str(tms_row['tracking_number']).lstrip("'")
                        tms_data[tracking_num] = tms_row
                        tms_data[f"'{tracking_num}"] = tms_row
                    logging.info(f"✓ Found {len(tms_results)} TMS records")
                    
                    # Now get ALL child bills for each master tracking number found
                    master_tracking_numbers = set()
                    for tms_row in tms_results:
                        master = tms_row.get('master_tracking_number')
                        if master:
                            master_tracking_numbers.add(master)
                    
                    if master_tracking_numbers:
                        logging.info(f"Fetching child bills for {len(master_tracking_numbers)} master tracking numbers")
                        # Query database for all bills with these master tracking numbers
                        import sqlite3
                        conn = sqlite3.connect(db_path)
                        cursor = conn.cursor()
                        placeholders = ','.join(['?' for _ in master_tracking_numbers])
                        query = f"SELECT tracking_number, master_tracking_number, customer_name, api_cost, charged_amount, weight_kg, created_at FROM tms_records WHERE master_tracking_number IN ({placeholders})"
                        cursor.execute(query, list(master_tracking_numbers))
                        all_children = cursor.fetchall()
                        conn.close()
                        
                        # Add any new children to tms_data
                        for row in all_children:
                            tracking_num = str(row[0]).lstrip("'")
                            if tracking_num not in tms_data:
                                tms_data[tracking_num] = {
                                    'tracking_number': row[0],
                                    'master_tracking_number': row[1],
                                    'customer_name': row[2],
                                    'api_cost': row[3],
                                    'charged_amount': row[4],
                                    'weight_kg': row[5],
                                    'created_at': row[6]
                                }
                        logging.info(f"✓ Total TMS records: {len(tms_data)}")
                else:
                    logging.warning("⚠ No TMS records found for tracking numbers")
            else:
                logging.warning("⚠ TMS database not found")
                
        except Exception as e:
            # If database query fails, continue without TMS data
            logging.error(f"TMS database query failed: {e}")
        
        # Phase 3: Build comparison with customer markups applied
        customer_pricing = load_customer_pricing()
        customers_dict = {c['name']: c for c in customer_pricing.get('customers', [])}
        
        # Process each result to build customer adjustment comparison
        tms_comparison = []
        customers_list = set()
        unfound_in_tms = []
        for result in results:
            tracking_id = result.get('tracking_id')
            fedex_master = str(result.get('TDMasterTrackingID', '')).lstrip("'")
            if not tracking_id:
                continue
            tracking_id_clean = str(tracking_id).lstrip("'")
            # Find TMS record for this tracking number
            tms_info = tms_data.get(tracking_id_clean)
            if not tms_info:
                unfound_in_tms.append(result)
                continue
            # Find all TMS records with the same master tracking number
            tms_master = tms_info.get('master_tracking_number', tracking_id_clean)
            tms_siblings = [rec for rec in tms_data.values() if rec.get('master_tracking_number', rec.get('tracking_number')) == tms_master]
            sibling_ids = set(rec.get('tracking_number') for rec in tms_siblings)
            sibling_count = len(sibling_ids) if sibling_ids else 1
            master_record = next((rec for rec in tms_siblings if rec.get('tracking_number') == tms_master), None)
            if not master_record:
                master_record = tms_siblings[0] if tms_siblings else tms_info
            total_api_cost = float(master_record.get('api_cost', 0) or 0)
            total_charged = float(master_record.get('charged_amount', 0) or 0)
            avg_api_cost = total_api_cost / sibling_count if sibling_count else total_api_cost
            avg_charged = total_charged / sibling_count if sibling_count else total_charged
            customer_name = tms_info.get('customer_name', '')
            customers_list.add(customer_name)


            # --- NEW: Use actual FedEx invoice amounts for customer adjustment ---
            # Find actual base rate and surcharges from invoice_surcharges and actual_base_rate
            invoice_surcharges = result.get('invoice_surcharges', [])
            actual_base_rate = 0.0
            surcharge_total = 0.0
            surcharge_lines = []
            # Find the actual base rate from breakdown (type == 'Base Rate')
            for item in result.get('breakdown', []):
                if item.get('type', '').lower() == 'base rate':
                    actual_base_rate = item.get('actual', 0.0)
            # Surcharges: sum all invoice_surcharges except base rate, performance pricing, grace discount, earned discount
            def is_excluded_surcharge(desc):
                desc_l = desc.lower()
                return (
                    'base rate' in desc_l or
                    ('performance' in desc_l and 'pricing' in desc_l) or
                    ('grace' in desc_l and 'discount' in desc_l) or
                    ('earned' in desc_l and 'discount' in desc_l)
                )
            for sc in invoice_surcharges:
                desc = sc.get('description', '')
                amount = sc.get('amount', 0.0)
                if is_excluded_surcharge(desc):
                    continue
                surcharge_total += amount
                # Show surcharge with markup for export report
                markup_pct = 0.0
                if customer_name in customers_dict:
                    customer = customers_dict[customer_name]
                    if 'fuel' in desc.lower():
                        markup_pct = customer.get('fuel_markup', 0)
                    elif any(kw in desc.lower() for kw in ['zone', 'extended', 'delivery area']):
                        markup_pct = customer.get('zone_markup', 0)
                    elif any(kw in desc.lower() for kw in ['demand', 'peak']):
                        markup_pct = customer.get('demand_markup', 0)
                    else:
                        markup_pct = customer.get('fixed_markup', 0)
                surcharge_with_markup = amount * (1 + markup_pct / 100)
                surcharge_lines.append(f"{desc} ${surcharge_with_markup:.2f}")

            # Apply customer markups to actual FedEx invoice values
            total_with_markup = 0.0
            base_rate_with_markup = 0.0
            surcharge_total_with_markup = 0.0
            if customer_name in customers_dict:
                customer = customers_dict[customer_name]
                # Base Rate markup
                base_markup = customer.get('base_markup', 0)
                base_rate_with_markup = actual_base_rate * (1 + base_markup / 100)
                total_with_markup += base_rate_with_markup
                # Surcharges markup
                for sc in invoice_surcharges:
                    desc = sc.get('description', '').lower()
                    amount = sc.get('amount', 0.0)
                    if is_excluded_surcharge(desc):
                        continue
                    if 'fuel' in desc:
                        markup_pct = customer.get('fuel_markup', 0)
                    elif any(kw in desc for kw in ['zone', 'extended', 'delivery area']):
                        markup_pct = customer.get('zone_markup', 0)
                    elif any(kw in desc for kw in ['demand', 'peak']):
                        markup_pct = customer.get('demand_markup', 0)
                    else:
                        markup_pct = customer.get('fixed_markup', 0)
                    surcharge_with_markup = amount * (1 + markup_pct / 100)
                    surcharge_total_with_markup += surcharge_with_markup
                    total_with_markup += surcharge_with_markup
            else:
                total_with_markup = None

            service_type = result.get('service_type', '')

            # Build comparison row
            comparison_row = {
                'tracking_id': tracking_id,
                'invoice_number': result.get('invoice_number', ''),
                'customer': customer_name,
                'zone': result.get('zone', ''),
                'rated_weight': result.get('rated_weight', 0),
                'actual_charge': result.get('actual_charge', 0),
                'expected_charge': result.get('expected_charge', 0),
                'difference': result.get('difference', 0),
                'status': result.get('status', ''),
                'ship_date': result.get('ship_date', ''),
                'dimensions': result.get('dimensions', ''),
                'ground_service': result.get('ground_service', ''),
                'pod_date': result.get('pod_date', ''),
                'actual_weight': result.get('actual_weight', 0),
                'tms_api_cost': avg_api_cost,
                'tms_charged': avg_charged,
                'expected_customer_charge': (round(total_with_markup, 2) if total_with_markup is not None else None),
                'customer_charge_diff': round((avg_charged or 0) - (total_with_markup or 0), 2) if total_with_markup is not None else None,
                'fedex_vs_api_diff': round(result.get('actual_charge', 0) - avg_api_cost, 2),
                'detailed_breakdown': result.get('breakdown', []),
                'TDMasterTrackingID': result.get('TDMasterTrackingID', ''),
                'service_type': service_type,
                'base_rate': round(base_rate_with_markup, 2) if total_with_markup is not None else None,
                'surcharge': "\n".join(surcharge_lines),
                'surcharge_total': round(surcharge_total_with_markup, 2) if total_with_markup is not None else None,
                'original_row': result.get('original_row', {}),
                'original_column_order': result.get('original_column_order', [])
            }
            tms_comparison.append(comparison_row)
        
        # Save each adjustment report shipment to the database (bulk insert for performance)
        logging.info(f"Phase 3: Saving {len(tms_comparison)} shipments to database")
        shipments_to_save = []
        
        for idx, comp in enumerate(tms_comparison):
            try:
                # Use TMS customer_name if available, else fallback
                tms_info = None
                tracking_id_clean = str(comp.get('tracking_id', '')).lstrip("'")
                if tracking_id_clean in tms_data:
                    tms_info = tms_data[tracking_id_clean]
                customer_name = ''
                if tms_info and tms_info.get('customer_name'):
                    customer_name = tms_info.get('customer_name')
                else:
                    customer_name = comp.get('customer', '')

                # Store ship_date in YYYY-MM-DD format for dashboard filtering
                raw_ship_date = comp.get('ship_date', '')
                ship_date_db = ''
                if raw_ship_date:
                    # Try to parse yyyymmdd or other formats
                    try:
                        if len(raw_ship_date) == 8 and raw_ship_date.isdigit():
                            # yyyymmdd
                            ship_date_db = f"{raw_ship_date[:4]}-{raw_ship_date[4:6]}-{raw_ship_date[6:]}"
                        else:
                            # Try parsing as date
                            dt = datetime.strptime(raw_ship_date, '%Y-%m-%d')
                            ship_date_db = dt.strftime('%Y-%m-%d')
                    except Exception:
                        ship_date_db = raw_ship_date

                # Store adjustment report values in DB
                shipment = {
                    'tracking_number': comp.get('tracking_id'),
                    'invoice_number': comp.get('invoice_number', ''),
                    'ship_date': ship_date_db,
                    'customer_name': customer_name,
                    'zone': comp.get('zone', ''),
                    'weight': comp.get('rated_weight', 0),
                    'service_type': comp.get('service_type', ''),
                    'actual_charge': comp.get('tms_charged', 0),  # Prepaid Charge
                    'expected_charge': comp.get('expected_customer_charge', 0),  # Total Amount
                    'adjustment': comp.get('customer_charge_diff', 0),  # Adjustment
                    'surcharge_breakdown': json.dumps(comp.get('detailed_breakdown', [])),
                    'surcharges_total': comp.get('surcharge_total', 0),
                    'verification_status': comp.get('status', ''),
                    'upload_timestamp': datetime.now(timezone.utc).isoformat(),
                    'analysis_result': json.dumps(comp)
                }
                shipments_to_save.append(shipment)
            except Exception as e:
                print(f"[FedEx DB] Error preparing shipment {comp.get('tracking_id')}: {e}")
        
        # Bulk insert all shipments in a single transaction (much faster for 10k+ rows)
        if shipments_to_save:
            try:
                saved_count = bulk_upsert_fedex_shipments(FEDEX_DB_PATH, shipments_to_save)
                logging.info(f"✓ Saved {saved_count} shipments to database")
            except Exception as e:
                logging.error(f"Bulk insert failed: {e}")
                # Fallback to individual inserts if bulk fails
                logging.info("Falling back to individual inserts...")
                for shipment in shipments_to_save:
                    try:
                        upsert_fedex_shipment(FEDEX_DB_PATH, shipment)
                    except Exception as e2:
                        logging.error(f"Error saving shipment {shipment.get('tracking_number')}: {e2}")
        
        # Store analysis results in cache for efficient export (avoid large POST payload)
        # Use current_user.id as cache key for reliable session matching
        cache_key = f"user_{current_user.id}"
        # Get original filename from JSON payload (sent by frontend)
        original_filename = invoice_data.get('original_filename') if invoice_data else None
        analysis_results_cache[cache_key] = {
            'tms_comparison': tms_comparison,
            'unfound_in_tms': unfound_in_tms,
            'original_filename': original_filename,
            'timestamp': datetime.now().isoformat()
        }
        logging.info(f"Cached {len(tms_comparison)} tms_comparison records for user {current_user.id}, filename: {original_filename}")
        
        return jsonify({
            'results': results,
            'tms_comparison': tms_comparison,
            'customers': sorted(list(customers_list)),
            'unfound_in_tms': unfound_in_tms
        })
    except Exception as e:
        logging.error(f"Verification error: {str(e)}", exc_info=True)
        return jsonify({'error': str(e)}), 500


@app.route('/api/upload_tms', methods=['POST'])
@login_required
def api_upload_tms():
    """Upload TMS Excel file and import into database"""
    try:
        # Check if file was uploaded
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'Empty filename'}), 400
        
        # Validate file extension
        if not file.filename.lower().endswith(('.xlsx', '.xls')):
            return jsonify({'error': 'Only Excel files (.xlsx, .xls) are supported'}), 400
        
        logging.info(f"[TMS Upload] Received file: {file.filename} from user {current_user.username}")
        
        # Save file temporarily
        import tempfile
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            file.save(tmp_file.name)
            tmp_path = tmp_file.name
        
        try:
            # Import the TMS export functions
            from tms_scraper import parse_tms_excel
            from tms_database import insert_records, get_database_stats
            
            # Parse the Excel file
            logging.info("[TMS Upload] Parsing Excel file...")
            all_records = parse_tms_excel(tmp_path)
            
            if not all_records:
                return jsonify({'error': 'No records found in Excel file'}), 400
            
            logging.info(f"[TMS Upload] ✓ Parsed {len(all_records)} records from Excel")
            
            # Insert all records into database
            logging.info(f"[TMS Upload] Inserting {len(all_records)} records into database...")
            inserted = insert_records(all_records)
            
            # Get database statistics
            stats = get_database_stats()
            
            logging.info("="*60)
            logging.info("[TMS Upload] ✓ IMPORT COMPLETED")
            logging.info(f"  Records parsed: {len(all_records)}")
            logging.info(f"  Records inserted: {inserted}")
            logging.info(f"  Total in database: {stats['total_records']}")
            logging.info(f"  Unique tracking #s: {stats['unique_tracking_numbers']}")
            logging.info("="*60)
            
            return jsonify({
                'success': True,
                'message': f'Successfully imported {inserted} records',
                'stats': {
                    'parsed': len(all_records),
                    'inserted': inserted,
                    'total_in_db': stats['total_records'],
                    'unique_tracking': stats['unique_tracking_numbers']
                }
            })
            
        finally:
            # Clean up temporary file
            try:
                os.unlink(tmp_path)
            except:
                pass
                
    except Exception as e:
        logging.error(f"[TMS Upload] Error: {str(e)}", exc_info=True)
        return jsonify({'error': str(e)}), 500


def calculate_customer_charge(verification_result, customer_pricing):
    """Calculate expected customer charge based on breakdown and markup percentages"""
    if not verification_result.get('breakdown'):
        return 0.0
    
    total = 0.0
    for item in verification_result['breakdown']:
        item_type = item.get('type', '').lower()
        expected = item.get('expected', 0)
        
        # Apply appropriate markup based on charge type
        if 'base rate' in item_type:
            markup = 1 + (customer_pricing.get('base_markup', 0) / 100)
        elif 'fuel' in item_type:
            markup = 1 + (customer_pricing.get('fuel_markup', 0) / 100)
        elif any(keyword in item_type for keyword in ['zone', 'extended']):
            markup = 1 + (customer_pricing.get('zone_markup', 0) / 100)
        elif any(keyword in item_type for keyword in ['demand', 'peak']):
            markup = 1 + (customer_pricing.get('demand_markup', 0) / 100)
        else:
            # Default to fixed surcharge markup
            markup = 1 + (customer_pricing.get('fixed_markup', 0) / 100)
        
        total += expected * markup
    
    return round(total, 2)


def verify_row(row, config):
    """Verify a single invoice row - config parameter is deprecated, now uses versioned rates"""
    from datetime import datetime
    from decimal import Decimal, ROUND_HALF_UP
    
    tracking_id = row.get('Express or Ground Tracking ID', '')
    invoice_number = row.get('Invoice Number', '')
    master_tracking_id = row.get('TDMasterTrackingID', '')
    
    # Safely parse numeric fields
    try:
        net_charge_str = row.get('Net Charge Amount', '')
        net_charge = float(net_charge_str) if net_charge_str and str(net_charge_str).strip() else 0.0
    except (ValueError, TypeError):
        net_charge = 0.0
    
    service_type = row.get('Service Type', '')
    shipment_date_str = row.get('Shipment Date', '')
    
    try:
        rated_weight_str = row.get('Rated Weight Amount', '')
        rated_weight = float(rated_weight_str) if rated_weight_str and str(rated_weight_str).strip() else 0.0
    except (ValueError, TypeError):
        rated_weight = 0.0
    
    zone_code = row.get('Zone Code', '')
    
    # Extract zone number from zone code (e.g., "Zone 2" -> 2)
    zone_num = None
    if zone_code:
        try:
            # Handle different formats: "Zone 2", "2", " 2 ", etc.
            zone_str = str(zone_code).replace('Zone', '').strip()
            zone_num = int(zone_str)
        except:
            zone_num = None
    
    # Parse shipment date
    shipment_date = None
    if shipment_date_str:
        try:
            # Try MM/DD/YYYY format
            shipment_date = datetime.strptime(shipment_date_str, '%m/%d/%Y')
        except:
            try:
                # Try YYYY-MM-DD format
                shipment_date = datetime.strptime(shipment_date_str, '%Y-%m-%d')
            except:
                try:
                    # Try YYYYMMDD format
                    shipment_date = datetime.strptime(str(shipment_date_str), '%Y%m%d')
                except:
                    pass
    
    # Get the appropriate rate version based on shipment date
    config = get_version_by_date(shipment_date)
    
    # Calculate expected charge
    expected_charge = 0
    breakdown = []
    charges_for_fuel = []  # Track individual charges for fuel calculation
    
    # Extract surcharges from invoice FIRST (before calculating base rate)
    invoice_surcharges = extract_invoice_surcharges(row)
    
    # Calculate FedEx's actual base rate from Transportation Charge + pricing adjustments
    def safe_float(val):
        try:
            return float(val) if val and str(val).strip() else 0.0
        except (ValueError, TypeError):
            return 0.0
    
    transportation_charge = safe_float(row.get('Transportation Charge Amount', 0))
    
    # Find Performance Pricing, Grace Discount, Earned Discount from invoice_surcharges
    performance_pricing = sum(isc['amount'] for isc in invoice_surcharges if 'performance' in isc['description'].lower() and 'pricing' in isc['description'].lower())
    grace_discount = sum(isc['amount'] for isc in invoice_surcharges if 'grace' in isc['description'].lower() and 'discount' in isc['description'].lower())
    earned_discount = sum(isc['amount'] for isc in invoice_surcharges if 'earned' in isc['description'].lower() and 'discount' in isc['description'].lower())
    
    actual_base_rate = transportation_charge + performance_pricing + grace_discount + earned_discount
    
    # 1. Base rate
    base_rate = get_base_rate(config, zone_num, rated_weight)
    if base_rate:
        expected_charge += base_rate
        breakdown.append({
            'type': 'Base Rate',
            'expected': base_rate,
            'actual': actual_base_rate,
            'difference': round(actual_base_rate - base_rate, 2)
        })
        charges_for_fuel.append(base_rate)
    
    # Check for Mixed Address case early
    residential_matches = [isc for isc in invoice_surcharges if 'residential' in isc['description'].lower()]
    is_ground = 'ground' in service_type.lower() and 'home delivery' not in service_type.lower()
    is_mixed_address = is_ground and len(residential_matches) > 0
    
    # 3. Verify fixed surcharges (exclude base rate components and residential for Ground service)
    for sc in config.get('fixed_surcharges', []):
        if sc.get('enabled', True):
            desc = sc.get('description', '')
            amount = float(sc.get('amount', 0))
            
            # Skip residential surcharge if this is Ground service (will be handled as Mixed Address)
            if 'residential' in desc.lower() and is_ground:
                continue
            
            # Check if this surcharge is in the invoice (but skip base rate components)
            matching = [isc for isc in invoice_surcharges 
                       if desc.lower() in isc['description'].lower()
                       and 'performance' not in isc['description'].lower()
                       and 'grace' not in isc['description'].lower()
                       and 'earned' not in isc['description'].lower()]
            if matching:
                actual_amount = sum(isc['amount'] for isc in matching)
                expected_charge += amount
                breakdown.append({
                    'type': desc,
                    'expected': amount,
                    'actual': actual_amount,
                    'difference': round(actual_amount - amount, 2)
                })
                charges_for_fuel.append(amount)
    
    # 4. Verify zone-based surcharges (exclude base rate components)
    for sc in config.get('zone_based_surcharges', []):
        if sc.get('enabled', True):
            desc = sc.get('description', '')
            # Check if this surcharge is in the invoice (but skip base rate components)
            matching = [isc for isc in invoice_surcharges 
                       if desc.lower() in isc['description'].lower()
                       and 'performance' not in isc['description'].lower()
                       and 'grace' not in isc['description'].lower()
                       and 'earned' not in isc['description'].lower()]
            if matching:
                amount = get_zone_based_amount(sc, zone_num)
                if amount:
                    actual_amount = sum(isc['amount'] for isc in matching)
                    expected_charge += amount
                    breakdown.append({
                        'type': desc,
                        'expected': amount,
                        'actual': actual_amount,
                        'difference': round(actual_amount - amount, 2)
                    })
                    charges_for_fuel.append(amount)
    
    # 5. Verify demand surcharges
    for sc in config.get('demand_surcharges', []):
        if sc.get('enabled', True) and shipment_date:
            desc = sc.get('description', '')
            service_filter = sc.get('service_type', '')
            
            # Check service type filter
            if service_filter and service_filter.lower() not in service_type.lower():
                continue
            
            # Check if date is in range
            amount = get_demand_surcharge_amount(sc, shipment_date)
            # Check if this surcharge is in the invoice (but skip base rate components)
            matching = [isc for isc in invoice_surcharges 
                       if desc.lower() in isc['description'].lower()
                       and 'performance' not in isc['description'].lower()
                       and 'grace' not in isc['description'].lower()
                       and 'earned' not in isc['description'].lower()]
            if amount and matching:
                actual_amount = sum(isc['amount'] for isc in matching)
                expected_charge += amount
                breakdown.append({
                    'type': desc,
                    'expected': amount,
                    'actual': actual_amount,
                    'difference': round(actual_amount - amount, 2)
                })
                charges_for_fuel.append(amount)
    
    # 6. Fuel surcharge - FedEx calculates item-by-item and rounds each
    if 'ground' in service_type.lower() and shipment_date:
        fuel_config = config.get('fuel_surcharge', {})
        fuel_pct = get_fuel_surcharge_percentage(fuel_config, shipment_date)
        if fuel_pct and fuel_pct > 0:
            # Calculate fuel on each charge individually using Decimal for precision
            fuel_amount = Decimal('0')
            fuel_pct_decimal = Decimal(str(fuel_pct)) / Decimal('100')
            
            for charge in charges_for_fuel:
                charge_decimal = Decimal(str(charge))
                item_fuel = (charge_decimal * fuel_pct_decimal).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
                fuel_amount += item_fuel
            
            fuel_amount = float(fuel_amount)
            
            # Get actual fuel from invoice
            fuel_matches = [isc for isc in invoice_surcharges if 'fuel' in isc['description'].lower()]
            actual_fuel = sum(isc['amount'] for isc in fuel_matches) if fuel_matches else 0
            
            expected_charge += fuel_amount
            breakdown.append({
                'type': f'Fuel Surcharge ({fuel_pct}%)',
                'expected': fuel_amount,
                'actual': actual_fuel,
                'difference': round(actual_fuel - fuel_amount, 2)
            })
    
    # Check for special cases before final comparison
    # 1. Check for MultiWeight (special discount we don't calculate)
    multiweight_weight = safe_float(row.get('Multiweight Total Multiweight Weight', 0))
    is_multiweight = multiweight_weight != 0.0
    
    # 2. Add Residential surcharge to breakdown for mixed address (already checked above)
    if is_mixed_address:
        residential_actual = sum(isc['amount'] for isc in residential_matches)
        breakdown.append({
            'type': 'Residential',
            'expected': 0.0,
            'actual': residential_actual,
            'difference': residential_actual
        })
    
    # Compare with actual
    difference = round(net_charge - expected_charge, 2)
    
    # Determine status
    if is_multiweight:
        status = 'MultiWeight'
    elif is_mixed_address:
        status = 'Mixed Address'
    elif abs(difference) < 0.01:
        status = 'match'
    else:
        status = 'mismatch'
    
    # Extract dimension fields
    dim_length = row.get('Dim Length', '')
    dim_width = row.get('Dim Width', '')
    dim_height = row.get('Dim Height', '')
    dimensions = f"{dim_length}x{dim_width}x{dim_height}" if dim_length and dim_width and dim_height else ''
    
    # Get ground service type
    ground_service = row.get('Ground Service', '')
    
    # Get POD delivery date
    pod_delivery_date_str = row.get('POD Delivery Date', '')
    pod_date = ''
    if pod_delivery_date_str:
        try:
            pod_dt = datetime.strptime(str(pod_delivery_date_str), '%m/%d/%Y')
            pod_date = pod_dt.strftime('%Y%m%d')
        except:
            try:
                pod_dt = datetime.strptime(str(pod_delivery_date_str), '%Y-%m-%d')
                pod_date = pod_dt.strftime('%Y%m%d')
            except:
                pod_date = str(pod_delivery_date_str)
    
    # Format shipment date
    ship_date_formatted = shipment_date.strftime('%Y%m%d') if shipment_date else ''
    
    return {
        'tracking_id': tracking_id,
        'TDMasterTrackingID': master_tracking_id,
        'invoice_number': invoice_number,
        'service_type': row.get('Service Type', service_type),
        'zone': zone_code,
        'rated_weight': rated_weight,
        'actual_charge': net_charge,
        'expected_charge': round(expected_charge, 2),
        'difference': difference,
        'status': status,
        'breakdown': breakdown,
        'invoice_surcharges': invoice_surcharges,
        'ship_date': ship_date_formatted,
        'dimensions': dimensions,
        'ground_service': ground_service,
        'pod_date': pod_date,
        'actual_weight': rated_weight,
        'charge': net_charge
    }


def get_base_rate(config, zone_num, weight):
    """Get base rate from rate table"""
    if not zone_num or zone_num < 2 or zone_num > 8:
        return None
    
    zone_idx = zone_num - 2  # Zone 2 is index 0
    weight_idx = int(weight) - 1  # 1 lb is index 0
    
    if weight_idx < 0 or weight_idx >= 150:
        return None
    
    try:
        rate = config['base_table']['rates'][weight_idx][zone_idx]
        return float(rate) if rate else 0
    except:
        return 0


def get_zone_based_amount(surcharge, zone_num):
    """Get zone-based surcharge amount"""
    if not zone_num:
        return None
    
    zone_rates = surcharge.get('zone_rates', {})
    # Try exact match first
    if str(zone_num) in zone_rates:
        return float(zone_rates[str(zone_num)])
    
    # Try zone ranges (e.g., "3-4")
    for zone_key, amount in zone_rates.items():
        if '-' in zone_key:
            try:
                start, end = map(int, zone_key.split('-'))
                if start <= zone_num <= end:
                    return float(amount)
            except:
                pass
    
    return None


def get_demand_surcharge_amount(surcharge, shipment_date):
    """Get demand surcharge amount based on date"""
    from datetime import datetime
    
    date_ranges = surcharge.get('date_ranges', [])
    for dr in date_ranges:
        try:
            start = datetime.strptime(dr['start_date'], '%Y-%m-%d')
            end = datetime.strptime(dr['end_date'], '%Y-%m-%d')
            if start <= shipment_date <= end:
                return float(dr['amount'])
        except:
            pass
    
    return None


def get_fuel_surcharge_percentage(fuel_config, shipment_date):
    """Get fuel surcharge percentage based on date"""
    from datetime import datetime
    
    date_ranges = fuel_config.get('date_ranges', [])
    for dr in date_ranges:
        try:
            start = datetime.strptime(dr['start_date'], '%Y-%m-%d')
            end = datetime.strptime(dr['end_date'], '%Y-%m-%d')
            if start <= shipment_date <= end:
                return float(dr['percentage'])
        except:
            pass
    
    return None


def extract_invoice_surcharges(row):
    """Extract surcharge description-amount pairs from invoice row"""
    surcharges = []
    
    # Look for Tracking ID Charge Description and Amount pairs
    for key in row.keys():
        if 'Tracking ID Charge Description' in key:
            desc = row.get(key, '').strip()
            if desc:
                # Find corresponding amount
                amount_key = key.replace('Description', 'Amount')
                amount_str = row.get(amount_key, '')
                try:
                    if amount_str and str(amount_str).strip():
                        amount_val = float(amount_str)
                        surcharges.append({'description': desc, 'amount': amount_val})
                except (ValueError, TypeError):
                    # Skip invalid amounts
                    pass
    
    return surcharges


# Existing export endpoint (synchronous)

# New async export endpoint
@app.route('/api/export_customer_adjustments', methods=['POST'])
@login_required
def export_customer_adjustments_async():
    """Start background export and return job ID.
    
    Uses cached analysis results from server-side to avoid large POST payloads.
    Falls back to POST data if cache is not available.
    """
    try:
        # Use current_user.id as cache key for reliable session matching
        cache_key = f"user_{current_user.id}"
        cached_data = analysis_results_cache.get(cache_key)
        
        # Try to get data from POST body (fallback for smaller datasets)
        data = request.json or {}
        
        # Prefer cached data for large datasets
        if cached_data and len(cached_data.get('tms_comparison', [])) > 0:
            tms_comparison = cached_data.get('tms_comparison', [])
            unfound_in_tms = cached_data.get('unfound_in_tms', [])
            original_filename = cached_data.get('original_filename')
            logging.info(f"[EXPORT] Using cached data: {len(tms_comparison)} records")
        else:
            tms_comparison = data.get('tms_comparison', [])
            unfound_in_tms = data.get('unfound_in_tms', [])
            original_filename = data.get('original_filename', None)
            logging.info(f"[EXPORT] Using POST data: {len(tms_comparison)} records")
        
        if not tms_comparison and not unfound_in_tms:
            return jsonify({'error': 'No data to export. Please run verification first.'}), 400
        
        # Load customer pricing for markup calculation
        customer_pricing = load_customer_pricing()
        customers_dict = {c['name']: c for c in customer_pricing.get('customers', [])}
        
        # Get user info for export tracking
        exported_by_id = current_user.id
        exported_by_username = current_user.username
        
        job_id = str(uuid.uuid4())
        export_jobs[job_id] = {'status': 'pending'}
        thread = threading.Thread(
            target=background_export_customer_adjustments, 
            args=(job_id, tms_comparison, unfound_in_tms, original_filename, customers_dict, exported_by_id, exported_by_username)
        )
        thread.start()
        return jsonify({'job_id': job_id})
    except Exception as e:
        logging.error(f"[EXPORT] Error starting background export: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

# Endpoint to check export status
@app.route('/api/export_status/<job_id>', methods=['GET'])
def export_status(job_id):
    job = export_jobs.get(job_id)
    if not job:
        return jsonify({'error': 'Job not found'}), 404
    return jsonify(job)

# Endpoint to download export file
@app.route('/api/download_export/<job_id>', methods=['GET'])
def download_export(job_id):
    job = export_jobs.get(job_id)
    print(f"[DOWNLOAD] Requested job_id: {job_id}")
    if not job:
        print(f"[DOWNLOAD] Job not found: {job_id}")
        return jsonify({'error': 'Job not found'}), 404
    if job.get('status') != 'done':
        print(f"[DOWNLOAD] Job not done yet: {job_id} (status: {job.get('status')})")
        return jsonify({'error': 'File not ready'}), 404
    if not os.path.exists(job.get('file', '')):
        print(f"[DOWNLOAD] File does not exist: {job.get('file', '')}")
        return jsonify({'error': 'File not ready'}), 404
    print(f"[DOWNLOAD] Sending file: {job.get('file', '')}")
    response = send_file(job['file'], as_attachment=True)
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response


@app.route('/api/exports', methods=['GET'])
def api_list_exports():
    """List all available export files"""
    try:
        metadata = load_exports_metadata()
        # Add additional info and check if files still exist
        enriched = []
        for item in metadata:
            file_path = item.get('file_path')
            if file_path and os.path.exists(file_path):
                item['exists'] = True
                item['download_filename'] = os.path.basename(file_path)
            else:
                item['exists'] = False
                item['download_filename'] = ''
            enriched.append(item)
        # Sort by timestamp descending (newest first)
        enriched.sort(key=lambda x: x['timestamp'], reverse=True)
        return jsonify({'exports': enriched})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/exports/<export_id>', methods=['GET'])
def api_download_export_by_id(export_id):
    """Download a specific export file by ID"""
    try:
        metadata = load_exports_metadata()
        export_item = next((item for item in metadata if item['id'] == export_id), None)
        if not export_item:
            return jsonify({'error': 'Export not found'}), 404
        file_path = export_item['file_path']
        if not os.path.exists(file_path):
            return jsonify({'error': 'File not found'}), 404
        return send_file(file_path, as_attachment=True, download_name=os.path.basename(file_path))
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/exports/<export_id>', methods=['DELETE'])
@login_required
def api_delete_export(export_id):
    """Delete a specific export file (only owner, manager, or admin)"""
    try:
        metadata = load_exports_metadata()
        export_item = next((item for item in metadata if item['id'] == export_id), None)
        if not export_item:
            return jsonify({'error': 'Export not found'}), 404
        
        # Check permission: only the exporter, manager, or admin can delete
        user_role = current_user.role if hasattr(current_user, 'role') else 'user'
        exported_by_id = export_item.get('exported_by_id')
        
        can_delete = (
            user_role in ['admin', 'manager'] or
            exported_by_id == current_user.id
        )
        
        if not can_delete:
            return jsonify({'error': 'Permission denied. Only the exporter, manager, or admin can delete this file.'}), 403
        
        # Delete the file
        file_path = export_item['file_path']
        if os.path.exists(file_path):
            os.remove(file_path)
        
        # Remove from metadata
        metadata = [item for item in metadata if item['id'] != export_id]
        save_exports_metadata(metadata)
        
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/export_mismatch', methods=['POST'])
@login_required
def api_export_mismatch():
    """Export mismatch data to Excel file on server"""
    try:
        data = request.json
        original_filename = data.get('original_filename', 'unknown')
        mismatch_data = data.get('mismatch_data', [])
        row_count = len(mismatch_data)
        
        if not mismatch_data:
            return jsonify({'error': 'No mismatch data to export'}), 400
        
        # Generate filename
        timestamp = datetime.now().strftime('%Y-%m-%d_%H%M%S')
        base_name = original_filename.replace('.xlsx', '').replace('.xls', '').replace('.csv', '')
        filename = os.path.join(EXPORTS_DIR, f'{timestamp}_{base_name}_mismatch.xlsx')
        
        # Create Excel file using openpyxl
        wb = Workbook()
        ws = wb.active
        ws.title = 'Mismatches'
        
        # Get column headers from first row of data
        if mismatch_data:
            headers = list(mismatch_data[0].keys())
            ws.append(headers)
            
            # Apply header styling
            header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
            header_font = Font(name='Arial', size=10, bold=True)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # Add data rows
            for row_data in mismatch_data:
                row_values = [row_data.get(h, '') for h in headers]
                ws.append(row_values)
            
            # Apply data styling
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.font = Font(name='Arial', size=10)
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Save file
        wb.save(filename)
        
        # Add metadata
        file_size = os.path.getsize(filename)
        add_export_metadata(
            original_filename=original_filename,
            export_type='mismatch',
            file_path=filename,
            file_size=file_size,
            row_count=row_count,
            exported_by_id=current_user.id,
            exported_by_username=current_user.username
        )
        
        return jsonify({'success': True, 'filename': os.path.basename(filename)})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/search_tracking/<tracking_number>', methods=['GET'])
def api_search_tracking(tracking_number):
    """Search for shipment by tracking number"""
    try:
        import sqlite3
        import os
        
        # Clean tracking number
        tracking_number = str(tracking_number).strip().lstrip("'")
        
        print(f"[SEARCH] Looking for tracking number: '{tracking_number}'")
        print(f"[SEARCH] Database path: {FEDEX_DB_PATH}")
        print(f"[SEARCH] Database exists: {os.path.exists(FEDEX_DB_PATH)}")
        
        # Search in FedEx database
        conn = sqlite3.connect(FEDEX_DB_PATH)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        # Try multiple search patterns to catch any format
        cursor.execute("""
            SELECT * FROM fedex_shipments 
            WHERE tracking_number = ? 
               OR tracking_number = ? 
               OR tracking_number LIKE ?
               OR tracking_number LIKE ?
        """, (tracking_number, f"'{tracking_number}", f"%{tracking_number}%", f"%{tracking_number}"))
        
        fedex_row = cursor.fetchone()
        
        # Debug: Count total records
        cursor.execute("SELECT COUNT(*) as count FROM fedex_shipments")
        total_count = cursor.fetchone()['count']
        print(f"[SEARCH] Total shipments in DB: {total_count}")
        
        # Debug: Sample a few tracking numbers
        cursor.execute("SELECT tracking_number FROM fedex_shipments LIMIT 5")
        sample_tracking = [row['tracking_number'] for row in cursor.fetchall()]
        print(f"[SEARCH] Sample tracking numbers: {sample_tracking}")
        
        conn.close()
        
        if not fedex_row:
            print(f"[SEARCH] Not found: '{tracking_number}'")
            return jsonify({'found': False, 'message': 'Tracking number not found in analyzed invoices'})
        
        print(f"[SEARCH] Found shipment! Tracking: {fedex_row['tracking_number']}")
        
        # Convert to dict
        fedex_data = dict(fedex_row)
        
        # Parse surcharge breakdown if exists
        if fedex_data.get('surcharge_breakdown'):
            try:
                fedex_data['surcharge_breakdown'] = json.loads(fedex_data['surcharge_breakdown'])
            except:
                pass
        
        if fedex_data.get('analysis_result'):
            try:
                fedex_data['analysis_result'] = json.loads(fedex_data['analysis_result'])
            except:
                pass
        
        # Search in TMS database
        tms_data = None
        try:
            from tms_database import get_records_by_tracking_numbers
            import os
            
            tms_db_path = os.path.join(os.path.dirname(__file__), 'tms_data.db')
            if os.path.exists(tms_db_path):
                tms_results = get_records_by_tracking_numbers([tracking_number])
                if tms_results:
                    tms_data = tms_results[0]
        except Exception as e:
            print(f"Error fetching TMS data: {e}")
        
        return jsonify({
            'found': True,
            'fedex_data': fedex_data,
            'tms_data': tms_data
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/dashboard/customers')
def api_dashboard_customers():
    customers = get_dashboard_customers(FEDEX_DB_PATH)
    return jsonify({'customers': customers})

@app.route('/api/dashboard/tms_customers')
def api_dashboard_tms_customers():
    from tms_database import get_tms_customers_fedex
    customers = get_tms_customers_fedex()
    return jsonify({'customers': customers})

@app.route('/api/dashboard/orders_by_day')
def api_dashboard_orders_by_day():
    customer = request.args.get('customer')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    data = get_tms_orders_by_day_fedex(start_date, end_date, customer)
    return jsonify({'orders_by_day': data})

@app.route('/api/dashboard/orders_by_weight')
def api_dashboard_orders_by_weight():
    customer = request.args.get('customer')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    data = get_dashboard_orders_by_weight(FEDEX_DB_PATH, customer, start_date, end_date)
    return jsonify({'orders_by_weight': data})

@app.route('/api/dashboard/orders_by_zone')
def api_dashboard_orders_by_zone():
    customer = request.args.get('customer')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    data = get_dashboard_orders_by_zone(FEDEX_DB_PATH, customer, start_date, end_date)
    return jsonify({'orders_by_zone': data})

@app.route('/api/dashboard/surcharge_stats')
def api_dashboard_surcharge_stats():
    customer = request.args.get('customer')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    data = get_dashboard_surcharge_stats(FEDEX_DB_PATH, customer, start_date, end_date)
    return jsonify({'surcharge_stats': data})

@app.route('/api/dashboard/adjustment_stats')
def api_dashboard_adjustment_stats():
    customer = request.args.get('customer')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    data = get_dashboard_adjustment_stats(FEDEX_DB_PATH, customer, start_date, end_date)
    return jsonify({'adjustment_stats': data})

# ===============================
# Supplier Management API
# ===============================

SUPPLIERS_FILE = os.path.join(os.path.dirname(__file__), 'data', 'suppliers.json')

def load_suppliers():
    """Load supplier data from JSON file"""
    if os.path.exists(SUPPLIERS_FILE):
        try:
            with open(SUPPLIERS_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print(f"Error loading suppliers: {e}")
            return {'suppliers': []}
    return {'suppliers': []}

def save_suppliers(data):
    """Save supplier data to JSON file"""
    try:
        os.makedirs(os.path.dirname(SUPPLIERS_FILE), exist_ok=True)
        with open(SUPPLIERS_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        return True
    except Exception as e:
        print(f"Error saving suppliers: {e}")
        return False

# ============================================
# Quote Pricing API (for Customer Quote feature)
# ============================================

QUOTE_PRICING_PATH = os.path.join(os.path.dirname(__file__), 'data', 'quote_pricing.json')

def load_quote_pricing():
    """Load quote pricing configuration for Customer Quote feature"""
    if os.path.exists(QUOTE_PRICING_PATH):
        try:
            with open(QUOTE_PRICING_PATH, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            logging.error(f"Error loading quote pricing: {e}")
            return {"version": "1.0", "fuelRates": {}, "carriers": {}}
    return {"version": "1.0", "fuelRates": {}, "carriers": {}}

def save_quote_pricing(data):
    """Save quote pricing configuration"""
    try:
        data['lastUpdated'] = datetime.now(timezone.utc).strftime('%Y-%m-%d')
        with open(QUOTE_PRICING_PATH, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        return True
    except Exception as e:
        logging.error(f"Error saving quote pricing: {e}")
        return False

@app.route('/api/quote-pricing', methods=['GET'])
def api_get_quote_pricing():
    """Get quote pricing configuration for Customer Quote"""
    data = load_quote_pricing()
    return jsonify(data)

@app.route('/api/quote-pricing', methods=['POST'])
def api_save_quote_pricing():
    """Save quote pricing configuration"""
    try:
        data = request.get_json()
        if save_quote_pricing(data):
            return jsonify({'success': True, 'message': 'Quote pricing saved successfully'})
        else:
            return jsonify({'success': False, 'error': 'Failed to save quote pricing'}), 500
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/quote-pricing/carrier/<carrier_id>', methods=['GET'])
def api_get_quote_carrier_pricing(carrier_id):
    """Get quote pricing for a specific carrier"""
    data = load_quote_pricing()
    carrier = data.get('carriers', {}).get(carrier_id)
    if carrier:
        return jsonify(carrier)
    return jsonify({'error': 'Carrier not found'}), 404

@app.route('/api/quote-pricing/carrier/<carrier_id>', methods=['POST'])
def api_save_quote_carrier_pricing(carrier_id):
    """Save quote pricing for a specific carrier"""
    try:
        carrier_data = request.get_json()
        data = load_quote_pricing()
        if 'carriers' not in data:
            data['carriers'] = {}
        data['carriers'][carrier_id] = carrier_data
        if save_quote_pricing(data):
            return jsonify({'success': True, 'message': f'Carrier {carrier_id} pricing saved'})
        else:
            return jsonify({'success': False, 'error': 'Failed to save'}), 500
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/quote-pricing/fuel-rates', methods=['POST'])
def api_save_quote_fuel_rates():
    """Save quote fuel rates"""
    try:
        fuel_data = request.get_json()
        data = load_quote_pricing()
        data['fuelRates'] = fuel_data
        if save_quote_pricing(data):
            return jsonify({'success': True, 'message': 'Fuel rates saved'})
        else:
            return jsonify({'success': False, 'error': 'Failed to save'}), 500
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/suppliers', methods=['GET'])
def api_get_suppliers():
    """Get all suppliers"""
    data = load_suppliers()
    return jsonify(data)

@app.route('/api/suppliers', methods=['POST'])
def api_save_suppliers():
    """Save all supplier data"""
    try:
        data = request.get_json()
        if save_suppliers(data):
            return jsonify({'success': True, 'message': 'Suppliers saved successfully'})
        else:
            return jsonify({'success': False, 'error': 'Failed to save suppliers'}), 500
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ===============================
# Zone Charts & Product Mapping API
# ===============================

ZONE_CHARTS_FILE = os.path.join(os.path.dirname(__file__), 'data', 'zone_charts.json')

# FedEx, FedEx OS, FedEx AHS, and SmartPost all share the same FedEx zone chart
ZONE_CHART_CARRIER_MAP = {
    'FedExAHS': 'FedEx',
    'FedExOS': 'FedEx',
    'SmartPost': 'FedEx',
}

def _zone_chart_carrier(carrier):
    """Normalize carrier name for zone chart lookup"""
    return ZONE_CHART_CARRIER_MAP.get(carrier, carrier)

def load_zone_charts():
    """Load zone chart data from JSON file"""
    if os.path.exists(ZONE_CHARTS_FILE):
        try:
            with open(ZONE_CHARTS_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print(f"Error loading zone charts: {e}")
    return {'zone_charts': [], 'product_mapping': {}}

def save_zone_charts(data):
    """Save zone chart data to JSON file"""
    try:
        os.makedirs(os.path.dirname(ZONE_CHARTS_FILE), exist_ok=True)
        with open(ZONE_CHARTS_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        return True
    except Exception as e:
        print(f"Error saving zone charts: {e}")
        return False

@app.route('/api/zone-charts', methods=['GET'])
def api_get_zone_charts():
    """Get all zone chart data including product mapping"""
    data = load_zone_charts()
    return jsonify(data)

@app.route('/api/zone-charts', methods=['POST'])
def api_save_zone_charts():
    """Save zone chart data"""
    try:
        data = request.get_json()
        if save_zone_charts(data):
            return jsonify({'success': True, 'message': 'Zone charts saved successfully'})
        else:
            return jsonify({'success': False, 'error': 'Failed to save zone charts'}), 500
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/zone-charts/lookup', methods=['POST'])
def api_zone_lookup():
    """Look up zone for a given carrier, ship-from zip, and ship-to zip"""
    try:
        req = request.get_json()
        carrier = req.get('carrier', '')
        ship_from = req.get('ship_from', '')
        ship_to = req.get('ship_to', '')
        
        data = load_zone_charts()
        zone_charts = data.get('zone_charts', [])
        
        # Find matching zone chart (normalize carrier for shared charts)
        zone = None
        ship_to_prefix = ship_to[:3] if len(ship_to) >= 3 else ship_to
        chart_carrier = _zone_chart_carrier(carrier)
        
        for chart in zone_charts:
            if chart.get('carrier') == chart_carrier and chart.get('ship_from_zip') == ship_from:
                for entry in chart.get('entries', []):
                    start = entry.get('start_zip', '')
                    end = entry.get('end_zip', '')
                    if start <= ship_to_prefix <= end:
                        zone = entry.get('zone', '')
                        break
                break
        
        return jsonify({'zone': zone, 'carrier': carrier, 'ship_from': ship_from, 'ship_to': ship_to})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ===============================
# Supplier Dashboard API (Other Carriers)
# ===============================

@app.route('/api/supplier-dashboard/product-names')
def api_supplier_product_names():
    """Get unique product names that are NOT 自营FedEx"""
    product_names = get_tms_supplier_product_names()
    return jsonify({'product_names': product_names})

@app.route('/api/supplier-dashboard/orders_by_day')
def api_supplier_orders_by_day():
    """Get orders by day for supplier products, grouped by product_name"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    product_name = request.args.get('product_name')
    data = get_tms_orders_by_day_supplier(start_date, end_date, product_name)
    return jsonify({'orders_by_day': data})


# ===============================
# Supplier Cost Verification API
# ===============================

@app.route('/api/supplier-verify', methods=['POST'])
def api_supplier_verify():
    """Verify supplier costs against TMS data"""
    try:
        req = request.get_json()
        start_date = req.get('start_date')
        end_date = req.get('end_date')
        product_name = req.get('product_name')
        
        if not product_name:
            return jsonify({'error': 'product_name is required'}), 400
        
        # Load zone charts and product mapping
        zone_data = load_zone_charts()
        product_mapping = zone_data.get('product_mapping', {})
        zone_charts = zone_data.get('zone_charts', [])
        
        # Get mapping for this product
        mapping = product_mapping.get(product_name)
        if not mapping:
            return jsonify({'error': f'No carrier/supplier mapping configured for product: {product_name}'}), 400
        
        carrier = mapping.get('carrier', '')
        supplier_name = mapping.get('supplier', '')
        service_type = mapping.get('service_type', 'hd')  # hd or ground
        default_ship_from = mapping.get('default_ship_from', '')  # fallback ship-from zip
        
        # Load supplier data
        supplier_data = load_suppliers()
        suppliers = supplier_data.get('suppliers', [])
        supplier = None
        for s in suppliers:
            if s.get('name') == supplier_name:
                supplier = s
                break
        
        if not supplier:
            return jsonify({'error': f'Supplier not found: {supplier_name}'}), 400
        
        # Load fuel rates from quote pricing for the current fuel %
        quote_pricing = load_quote_pricing()
        fuel_rates = quote_pricing.get('fuelRates', {})
        
        # Get TMS records for this product
        records = get_tms_supplier_records_for_verification(start_date, end_date, product_name)
        
        results = []
        match_count = 0
        mismatch_count = 0
        no_zone_count = 0
        
        for record in records:
            ship_to_zip = record.get('ship_to_zip', '')
            ship_from_zip = record.get('ship_from_zip', '')
            
            # If ship_from_zip is missing, use the default from product mapping
            if not ship_from_zip and default_ship_from:
                ship_from_zip = default_ship_from
            
            weight_kg = record.get('weight_kg', 0)
            cargo_weight_kg = record.get('cargo_weight_kg', 0) or weight_kg  # 货物重量; fallback to 计费重
            tms_cost = record.get('api_cost', 0)
            tracking = record.get('tracking_number', '')
            customer_name = record.get('customer_name', '')
            tms_date = record.get('tms_created_at', '')
            tms_surcharges = record.get('tms_surcharges', {})
            
            # Use cargo weight (货物重量) for rate lookup — TMS prices based on actual weight
            # Keep billed weight (计费重) for display
            import math
            weight_lb = cargo_weight_kg * 2.20462
            weight_lb_rounded = max(1, math.ceil(weight_lb))  # Shipping: always round up
            
            # Look up zone (normalize carrier for shared zone charts)
            zone = None
            ship_to_prefix = ship_to_zip[:3] if len(ship_to_zip) >= 3 else ship_to_zip
            chart_carrier = _zone_chart_carrier(carrier)
            
            for chart in zone_charts:
                if chart.get('carrier') == chart_carrier and chart.get('ship_from_zip') == ship_from_zip:
                    for entry in chart.get('entries', []):
                        start = entry.get('start_zip', '')
                        end = entry.get('end_zip', '')
                        if start <= ship_to_prefix <= end:
                            zone = entry.get('zone', '')
                            break
                    break
            
            if not zone:
                no_zone_count += 1
                results.append({
                    'tracking_number': tracking,
                    'customer_name': customer_name,
                    'date': tms_date,
                    'ship_from': ship_from_zip,
                    'ship_to': ship_to_zip,
                    'weight_kg': weight_kg,
                    'weight_lb': round(weight_lb, 2),
                    'zone': None,
                    'tms_cost': tms_cost,
                    'calculated_cost': None,
                    'difference': None,
                    'status': 'no_zone',
                    'breakdown': {},
                    'tms_surcharges': tms_surcharges
                })
                continue
            
            # Calculate expected cost from supplier rates
            calculated = calculate_supplier_cost(supplier, zone, weight_lb_rounded, cargo_weight_kg, carrier, fuel_rates, service_type)
            
            difference = round(tms_cost - calculated['total'], 2) if calculated['total'] else None
            
            # Determine match status (within $0.02 tolerance)
            if difference is not None and abs(difference) <= 0.02:
                status = 'match'
                match_count += 1
            else:
                status = 'mismatch'
                mismatch_count += 1
            
            results.append({
                'tracking_number': tracking,
                'customer_name': customer_name,
                'date': tms_date,
                'ship_from': ship_from_zip,
                'ship_to': ship_to_zip,
                'weight_kg': weight_kg,
                'weight_lb': round(weight_lb, 2),
                'zone': zone,
                'tms_cost': tms_cost,
                'calculated_cost': calculated['total'],
                'difference': difference,
                'status': status,
                'breakdown': calculated['breakdown'],
                'tms_surcharges': tms_surcharges
            })
        
        return jsonify({
            'success': True,
            'product_name': product_name,
            'carrier': carrier,
            'supplier': supplier_name,
            'total': len(results),
            'match_count': match_count,
            'mismatch_count': mismatch_count,
            'no_zone_count': no_zone_count,
            'results': results
        })
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


def calculate_supplier_cost(supplier, zone, weight_lb, weight_kg, carrier, fuel_rates, service_type='hd'):
    """
    Calculate expected cost based on supplier rate tables.
    Returns dict with total and breakdown.
    """
    breakdown = {}
    total = 0
    
    # Determine which rate table to use based on weight
    # Suppliers may have both oz and lb tables; oz is for light packages,
    # lb is for heavier ones (e.g. USPS oz table covers 1-16 oz, lb covers 1-70 lb).
    rate_table = None
    weight_unit = 'lb'
    
    if supplier.get('base_rates_oz'):
        # Check if weight falls within the oz table range
        import math
        weight_oz = weight_kg * 35.274
        weight_oz_rounded = max(1, math.ceil(weight_oz))
        weight_end_oz = supplier.get('weight_end_oz', len(supplier['base_rates_oz']))
        
        if weight_oz_rounded <= weight_end_oz:
            rate_table = supplier['base_rates_oz']
            weight_unit = 'oz'
        elif supplier.get('hd_base_rates_lb'):
            # Weight exceeds oz range — fall back to lb table
            rate_table = supplier['hd_base_rates_lb']
            weight_unit = 'lb'
        else:
            # No lb table, cap at max oz row
            rate_table = supplier['base_rates_oz']
            weight_unit = 'oz'
    
    if rate_table is None:
        if service_type == 'ground' and supplier.get('ground_base_rates_lb'):
            rate_table = supplier['ground_base_rates_lb']
        elif supplier.get('hd_base_rates_lb'):
            rate_table = supplier['hd_base_rates_lb']
    
    # Pick the zones list that matches the rate table
    if weight_unit == 'oz':
        zones = supplier.get('zones_oz') or supplier.get('zones', [2, 3, 4, 5, 6, 7, 8])
    else:
        zones = supplier.get('hd_zones_lb') or supplier.get('zones', [2, 3, 4, 5, 6, 7, 8])
    
    # Try to find zone index
    try:
        zone_val = int(zone) if zone else None
    except (ValueError, TypeError):
        # Handle zone strings like "8+9"
        zone_val = None
        for i, z in enumerate(zones):
            if str(z) == str(zone):
                zone_val = z
                break
    
    zone_idx = None
    if zone_val is not None:
        for i, z in enumerate(zones):
            if str(z) == str(zone_val) or str(z) == str(zone):
                zone_idx = i
                break
    
    if zone_idx is None:
        return {'total': None, 'breakdown': {'error': f'Zone {zone} not found in supplier zones {zones}'}}
    
    if not rate_table:
        return {'total': None, 'breakdown': {'error': 'No rate table found for supplier'}}
    
    # Look up base rate
    base_rate = 0
    if weight_unit == 'oz':
        # For oz-based rates, convert kg to oz (1 kg = 35.274 oz)
        import math
        weight_oz = weight_kg * 35.274
        weight_oz_rounded = max(1, math.ceil(weight_oz))
        row_idx = min(weight_oz_rounded - 1, len(rate_table) - 1)
        if row_idx >= 0 and row_idx < len(rate_table):
            row = rate_table[row_idx]
            if zone_idx < len(row):
                base_rate = row[zone_idx] or 0
    else:
        # lb-based rates
        row_idx = min(weight_lb - 1, len(rate_table) - 1)
        if row_idx >= 0 and row_idx < len(rate_table):
            row = rate_table[row_idx]
            if zone_idx < len(row):
                base_rate = row[zone_idx] or 0
    
    breakdown['base_rate'] = round(base_rate, 2)
    total += base_rate
    
    # Fixed surcharges (always applied for now - residential is common)
    fixed_surcharges = supplier.get('fixed_surcharges', [])
    fixed_total = 0
    for surcharge in fixed_surcharges:
        desc = surcharge.get('description', '')
        amount = surcharge.get('amount', 0) or 0
        # Only auto-apply Residential surcharge; others need specific logic
        if desc.lower() in ('residential',):
            fixed_total += amount
            breakdown[f'surcharge_{desc}'] = round(amount, 2)
    
    total += fixed_total
    
    # Zone-based surcharges - skip for basic calculation (these require address analysis)
    # They would need DAS/remote flags from TMS data
    
    # Fuel surcharge
    fuel_amount = 0
    fuel_type = supplier.get('fuel_type', 'discount')
    no_fuel_hd = supplier.get('no_fuel_home_delivery', False)
    
    if not (no_fuel_hd and service_type == 'hd'):
        if fuel_type == 'fixed':
            fuel_pct = supplier.get('fuel_fixed', 0) or 0
        else:
            # Discount type: official fuel rate minus discount
            official_rate = 0
            fedex_fuel_carriers = ['FedEx', 'FedExAHS', 'FedExOS', 'SmartPost']
            ups_fuel_carriers = ['UPS']
            
            if carrier in fedex_fuel_carriers:
                official_rate = fuel_rates.get('fedex', 0) or 0
            elif carrier in ups_fuel_carriers:
                official_rate = fuel_rates.get('ups', 0) or 0
            
            discount = supplier.get('fuel_discount', 0) or 0
            fuel_pct = official_rate - discount
        
        if fuel_pct > 0:
            fuel_amount = base_rate * (fuel_pct / 100)
            breakdown['fuel_surcharge'] = round(fuel_amount, 2)
            breakdown['fuel_pct'] = fuel_pct
    
    total += fuel_amount
    
    # Pickup fee for UNIUNI
    if carrier == 'UNIUNI':
        pickup_fee = supplier.get('pickup_fee', 0) or 0
        if pickup_fee > 0:
            breakdown['pickup_fee'] = round(pickup_fee, 2)
            total += pickup_fee
    
    # Apply markup percentage (e.g. 0.5 means 0.5%)
    markup_pct = supplier.get('markup', 0) or 0
    if markup_pct > 0:
        markup_amount = total * (markup_pct / 100)
        breakdown['markup_pct'] = markup_pct
        breakdown['markup_amount'] = round(markup_amount, 2)
        total += markup_amount
    
    breakdown['total'] = round(total, 2)
    
    return {'total': round(total, 2), 'breakdown': breakdown}


# ── Scheduler API endpoints ─────────────────────────────────────────────

@app.route('/api/scheduler/status')
@login_required
def api_scheduler_status():
    """Return background scheduler status + job history."""
    if not current_user.permissions.get('admin'):
        return jsonify({"error": "Admin access required"}), 403
    return jsonify(get_scheduler_status())


@app.route('/api/scheduler/trigger/<job_id>', methods=['POST'])
@login_required
def api_scheduler_trigger(job_id):
    """Manually trigger a background job (wms_monitor or tms_export)."""
    if not current_user.permissions.get('admin'):
        return jsonify({"error": "Admin access required"}), 403
    ok = trigger_job_now(job_id)
    if ok:
        return jsonify({"success": True, "message": f"Job '{job_id}' triggered"})
    return jsonify({"error": f"Job '{job_id}' not found"}), 404


if __name__ == '__main__':
    # Allow overriding host/port via environment variables to avoid port conflicts
    host = os.environ.get('FLASK_HOST', '127.0.0.1')
    port = int(os.environ.get('PORT', os.environ.get('FLASK_RUN_PORT', 5000)))
    debug = os.environ.get('FLASK_DEBUG', '1') not in ('0', 'false', 'False')
    # Initialize databases on startup
    init_fedex_db(FEDEX_DB_PATH)
    init_user_db()
    # Create default admin user if no users exist
    create_default_admin(bcrypt)

    # Guard: only start background services once (Flask reloader spawns a child process)
    is_reloader_child = os.environ.get('WERKZEUG_RUN_MAIN') == 'true'
    if not debug or is_reloader_child:
        # Start the turnover FastAPI backend as a subprocess
        from turnover_blueprint import start_turnover_backend, stop_turnover_backend
        start_turnover_backend()
        # Start unified background scheduler (WMS monitor + TMS daily export)
        start_scheduler()
        import atexit
        atexit.register(stop_turnover_backend)
        atexit.register(stop_scheduler)

    app.run(debug=debug, host=host, port=port)
