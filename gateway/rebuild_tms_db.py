"""
TMS Database Rebuild Script
Imports complete TMS export into an optimized SQLite database.

Usage: python rebuild_tms_db.py [path_to_xlsx]

Defaults to: 订单列表导出_1771461216624.xlsx
"""
import sqlite3
import openpyxl
import json
import os
import sys
import time
import shutil
from datetime import datetime, date

DB_PATH = os.path.join(os.path.dirname(__file__), 'tms_data.db')
BACKUP_DIR = os.path.join(os.path.dirname(__file__), 'backups')

# Address fields that appear twice in TMS export (recipient then sender)
ADDRESS_FIELDS = {'姓名', '公司名称', '国家简码', '省、州', '城市', '区',
                  '街道地址1', '街道地址2', '街道地址3', '邮编', '电话', '邮箱', '门牌号'}


def disambiguate_headers(raw_headers):
    """
    Resolve duplicate column names in TMS export.
    1st address block → 收件人_X, 2nd → 发件人_X
    1st 计费重 → 计费重 (order), 2nd → 箱计费重 (box)
    """
    headers = []
    seen_counts = {}
    for h in raw_headers:
        if h is None:
            headers.append(None)
            continue
        seen_counts[h] = seen_counts.get(h, 0) + 1
        count = seen_counts[h]
        if count == 1:
            headers.append(h)
        elif count == 2:
            if h in ADDRESS_FIELDS:
                headers.append(f'发件人_{h}')
            elif h == '计费重':
                headers.append('箱计费重')
            else:
                headers.append(f'{h}_2')
        else:
            headers.append(f'{h}_{count}')

    # 2nd pass: rename 1st occurrence of address fields to 收件人_ if duplicates exist
    for h in ADDRESS_FIELDS:
        if seen_counts.get(h, 0) >= 2:
            for i, hdr in enumerate(headers):
                if hdr == h:
                    headers[i] = f'收件人_{h}'
                    break
    return headers


def safe_float(v):
    """Safely convert a value to float, returning 0.0 on failure."""
    if v is None or v == '' or v == 'None':
        return 0.0
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0.0


def safe_str(v):
    """Safely convert a value to string."""
    if v is None:
        return ''
    return str(v).strip()


def create_new_schema(conn):
    """Create the optimized tms_records table with proper columns."""
    cursor = conn.cursor()

    cursor.execute('DROP TABLE IF EXISTS tms_records')

    cursor.execute('''
        CREATE TABLE tms_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tracking_number TEXT NOT NULL UNIQUE,
            tms_order_number TEXT,
            master_tracking_number TEXT,
            customer_order_number TEXT,
            transfer_number TEXT,
            customer_name TEXT,
            product_name TEXT,
            order_status TEXT,
            tracking_status TEXT,
            settlement_status TEXT,
            
            api_cost REAL DEFAULT 0,
            charged_amount REAL DEFAULT 0,
            order_amount REAL DEFAULT 0,
            profit REAL DEFAULT 0,
            
            weight_kg REAL DEFAULT 0,
            cargo_weight_kg REAL DEFAULT 0,
            box_count INTEGER DEFAULT 1,
            
            address_type TEXT DEFAULT '',
            remote_type TEXT DEFAULT '',
            
            ship_to_zip TEXT DEFAULT '',
            ship_to_state TEXT DEFAULT '',
            ship_to_city TEXT DEFAULT '',
            ship_from_zip TEXT DEFAULT '',
            ship_from_state TEXT DEFAULT '',
            
            carrier_name TEXT DEFAULT '',
            original_carrier_name TEXT DEFAULT '',
            channel_name TEXT DEFAULT '',
            channel_code TEXT DEFAULT '',
            
            signature_service TEXT DEFAULT '',
            
            -- Payable surcharges (应付) from TMS
            pay_freight REAL DEFAULT 0,
            pay_fuel REAL DEFAULT 0,
            pay_residential REAL DEFAULT 0,
            pay_residential_addr REAL DEFAULT 0,
            pay_das REAL DEFAULT 0,
            pay_das_remote REAL DEFAULT 0,
            pay_das_remote_extreme REAL DEFAULT 0,
            pay_ahs REAL DEFAULT 0,
            pay_oversize REAL DEFAULT 0,
            pay_peak REAL DEFAULT 0,
            pay_peak_oversize REAL DEFAULT 0,
            pay_peak_residential REAL DEFAULT 0,
            pay_remote REAL DEFAULT 0,
            pay_extreme_remote REAL DEFAULT 0,
            pay_signature REAL DEFAULT 0,
            pay_address_change REAL DEFAULT 0,
            pay_other REAL DEFAULT 0,
            
            -- Receivable surcharges (应收) from TMS
            recv_freight REAL DEFAULT 0,
            recv_fuel REAL DEFAULT 0,
            recv_residential REAL DEFAULT 0,
            recv_das REAL DEFAULT 0,
            recv_ahs REAL DEFAULT 0,
            recv_other REAL DEFAULT 0,
            
            tms_created_at TIMESTAMP,
            charge_time TIMESTAMP,
            
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    # Indexes
    cursor.execute('CREATE UNIQUE INDEX idx_tracking ON tms_records(tracking_number)')
    cursor.execute('CREATE INDEX idx_product ON tms_records(product_name)')
    cursor.execute('CREATE INDEX idx_tms_date ON tms_records(tms_created_at)')
    cursor.execute('CREATE INDEX idx_customer ON tms_records(customer_name)')
    cursor.execute('CREATE INDEX idx_status ON tms_records(order_status)')
    cursor.execute('CREATE INDEX idx_product_date ON tms_records(product_name, tms_created_at)')

    conn.commit()
    print("[SCHEMA] Created optimized tms_records table with indexes")


def import_xlsx(file_path, conn):
    """Import Excel file into the new database schema."""
    print(f"[IMPORT] Loading workbook: {file_path}")
    t0 = time.time()
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    print(f"[IMPORT] Workbook loaded in {time.time()-t0:.1f}s")

    # Read and disambiguate headers
    raw_headers = [cell.value for cell in ws[1]]
    headers = disambiguate_headers(raw_headers)
    col = {h: i for i, h in enumerate(headers) if h}
    print(f"[IMPORT] {len(headers)} columns, disambiguated duplicates")

    # Map column indices
    def ci(name):
        return col.get(name)

    tracking_col = ci('子单号')
    order_num_col = ci('运单号')
    if tracking_col is None and order_num_col is None:
        raise Exception("Required column '子单号' or '运单号' not found!")

    status_col = ci('订单状态')
    if status_col is None:
        raise Exception("Required column '订单状态' not found!")

    cursor = conn.cursor()
    batch = []
    batch_size = 5000
    total_rows = 0
    skipped = 0
    inserted = 0

    print("[IMPORT] Reading rows...")

    for row in ws.iter_rows(min_row=2, values_only=True):
        total_rows += 1
        if total_rows % 50000 == 0:
            print(f"  ...processed {total_rows:,} rows ({inserted:,} inserted, {skipped:,} skipped)")

        if not row:
            skipped += 1
            continue

        # Filter: only 订单状态 = 已生成
        order_status = safe_str(row[status_col]) if status_col is not None else ''
        if order_status != '已生成':
            skipped += 1
            continue

        # Use 子单号 as tracking; fallback to 运单号
        tracking = safe_str(row[tracking_col]) if tracking_col is not None else ''
        tms_order_num = safe_str(row[order_num_col]) if order_num_col is not None else ''
        if not tracking:
            tracking = tms_order_num
        if not tracking:
            skipped += 1
            continue

        def _col_str(name):
            idx = ci(name)
            return safe_str(row[idx]) if idx is not None and idx < len(row) else ''

        def _col_float(name):
            idx = ci(name)
            return safe_float(row[idx]) if idx is not None and idx < len(row) else 0.0

        # Parse tms_created_at
        tms_created_at = None
        created_str = _col_str('创建时间')
        if created_str:
            try:
                tms_created_at = datetime.strptime(created_str, '%Y-%m-%d %H:%M:%S')
            except:
                try:
                    tms_created_at = datetime.strptime(created_str[:19], '%Y-%m-%dT%H:%M:%S')
                except:
                    pass

        charge_time = None
        charge_str = _col_str('扣款时间')
        if charge_str:
            try:
                charge_time = datetime.strptime(charge_str, '%Y-%m-%d %H:%M:%S')
            except:
                pass

        # 计费重 = billed weight (may include dimensional), 货物重量 = actual cargo weight
        weight = _col_float('计费重')
        cargo_weight = _col_float('货物重量')

        record = (
            tracking,                                          # tracking_number
            tms_order_num,                                      # tms_order_number
            _col_str('服务商单号'),                              # master_tracking_number
            _col_str('客户单号'),                                # customer_order_number
            _col_str('转单号'),                                  # transfer_number
            _col_str('商户名称'),                                # customer_name
            _col_str('产品名称'),                                # product_name
            order_status,                                       # order_status
            _col_str('轨迹状态'),                                # tracking_status
            _col_str('结算状态'),                                # settlement_status
            _col_float('应付金额'),                              # api_cost
            _col_float('扣款金额'),                              # charged_amount
            _col_float('订单金额'),                              # order_amount
            _col_float('利润'),                                  # profit
            weight,                                             # weight_kg (计费重)
            cargo_weight,                                       # cargo_weight_kg (货物重量)
            int(safe_float(_col_str('箱子数量')) or 1),          # box_count
            _col_str('地址类型'),                                # address_type
            _col_str('偏远类型'),                                # remote_type
            _col_str('收件人_邮编'),                              # ship_to_zip
            _col_str('收件人_省、州'),                            # ship_to_state
            _col_str('收件人_城市'),                              # ship_to_city
            _col_str('发件人_邮编'),                              # ship_from_zip
            _col_str('发件人_省、州'),                            # ship_from_state
            _col_str('服务商名称'),                              # carrier_name
            _col_str('原服务商名称'),                            # original_carrier_name
            _col_str('渠道名称'),                                # channel_name
            _col_str('渠道代码'),                                # channel_code
            _col_str('签名服务'),                                # signature_service
            # Payable surcharges
            _col_float('应付运费'),
            _col_float('应付燃油附加费'),
            _col_float('应付住宅附加费'),
            _col_float('应付住宅地址费'),
            _col_float('应付DAS'),
            _col_float('应付DAS Remote'),
            _col_float('应付DAS Remote 荒远地区'),
            _col_float('应付AHS'),
            _col_float('应付oversize附加费'),
            _col_float('应付旺季附加费'),
            _col_float('应付超长旺季附加费'),
            _col_float('应付住宅地址旺季附加费'),
            _col_float('应付偏远费') + _col_float('应付非常偏远费') + _col_float('应付超偏远费'),
            _col_float('应付商业地址极度偏远费') + _col_float('应付住宅地址极度偏远费'),
            _col_float('应付签名签收'),
            _col_float('应付更改地址费'),
            _col_float('应付其他'),
            # Receivable surcharges (summarized)
            _col_float('应收运费'),
            _col_float('应收燃油附加费'),
            _col_float('应收住宅附加费') + _col_float('应收住宅地址费'),
            _col_float('应收DAS') + _col_float('应收DAS Remote'),
            _col_float('应收AHS'),
            _col_float('应收其他'),
            # Timestamps
            tms_created_at,
            charge_time,
        )

        batch.append(record)

        if len(batch) >= batch_size:
            _insert_batch(cursor, batch)
            inserted += len(batch)
            batch = []

    # Insert remaining
    if batch:
        _insert_batch(cursor, batch)
        inserted += len(batch)

    conn.commit()
    wb.close()

    print(f"\n[IMPORT] Done!")
    print(f"  Total rows read: {total_rows:,}")
    print(f"  Inserted: {inserted:,}")
    print(f"  Skipped: {skipped:,}")
    print(f"  Time: {time.time()-t0:.1f}s")

    return inserted


def _insert_batch(cursor, batch):
    """Bulk insert a batch of records."""
    cursor.executemany('''
        INSERT OR REPLACE INTO tms_records (
            tracking_number, tms_order_number, master_tracking_number, customer_order_number, transfer_number,
            customer_name, product_name, order_status, tracking_status, settlement_status,
            api_cost, charged_amount, order_amount, profit,
            weight_kg, cargo_weight_kg, box_count,
            address_type, remote_type,
            ship_to_zip, ship_to_state, ship_to_city,
            ship_from_zip, ship_from_state,
            carrier_name, original_carrier_name, channel_name, channel_code,
            signature_service,
            pay_freight, pay_fuel, pay_residential, pay_residential_addr,
            pay_das, pay_das_remote, pay_das_remote_extreme,
            pay_ahs, pay_oversize,
            pay_peak, pay_peak_oversize, pay_peak_residential,
            pay_remote, pay_extreme_remote,
            pay_signature, pay_address_change, pay_other,
            recv_freight, recv_fuel, recv_residential, recv_das, recv_ahs, recv_other,
            tms_created_at, charge_time
        ) VALUES (
            ?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,
            ?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?
        )
    ''', batch)


def rebuild_database(xlsx_path=None):
    """Main entry point: backup old DB, create new one, import data."""

    if xlsx_path is None:
        xlsx_path = os.path.join(os.path.dirname(__file__), '订单列表导出_1771549656398.xlsx')

    if not os.path.exists(xlsx_path):
        print(f"[ERROR] File not found: {xlsx_path}")
        return False

    # Backup existing DB
    if os.path.exists(DB_PATH):
        os.makedirs(BACKUP_DIR, exist_ok=True)
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_path = os.path.join(BACKUP_DIR, f'tms_data_{ts}.db')
        print(f"[BACKUP] Backing up current DB to {backup_path}")
        shutil.copy2(DB_PATH, backup_path)
        print(f"[BACKUP] Done ({os.path.getsize(backup_path)/1024/1024:.1f} MB)")

    # Create fresh database
    print(f"\n[DB] Creating new database at {DB_PATH}")
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")
    conn.execute("PRAGMA cache_size=-64000")  # 64MB cache

    create_new_schema(conn)
    inserted = import_xlsx(xlsx_path, conn)

    # Analyze for query optimizer
    print("[DB] Running ANALYZE...")
    conn.execute("ANALYZE")
    conn.commit()

    # Show stats
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) FROM tms_records")
    total = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(DISTINCT product_name) FROM tms_records")
    products = cursor.fetchone()[0]
    cursor.execute("SELECT MIN(tms_created_at), MAX(tms_created_at) FROM tms_records")
    mn, mx = cursor.fetchone()

    conn.close()

    db_size = os.path.getsize(DB_PATH) / 1024 / 1024

    print(f"\n{'='*50}")
    print(f"  Database rebuilt successfully!")
    print(f"  Records: {total:,}")
    print(f"  Products: {products}")
    print(f"  Date range: {mn} to {mx}")
    print(f"  DB size: {db_size:.1f} MB")
    print(f"{'='*50}")

    return True


if __name__ == '__main__':
    xlsx = sys.argv[1] if len(sys.argv) > 1 else None
    rebuild_database(xlsx)
