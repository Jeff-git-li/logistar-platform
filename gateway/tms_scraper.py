#!/usr/bin/env python3
"""
TMS Scraper: Fetch tracking data from Eccang TMS system

This script automates login and data extraction from the TMS portal.
Configure credentials in .env file: TMS_USER and TMS_PASS
"""

import os
import time
import logging
from pathlib import Path
from typing import List, Dict, Optional

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl


ENV_PATH = Path(__file__).parent / ".env"
_ENV_LOADED = False


def load_env_file():
    """Load environment variables from .env file"""
    global _ENV_LOADED
    if _ENV_LOADED:
        return
    
    try:
        if ENV_PATH.exists():
            with open(ENV_PATH, "r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if not line or line.startswith("#"):
                        continue
                    if "=" in line:
                        key, value = line.split("=", 1)
                        os.environ[key.strip()] = value.strip()
            logging.info(f"Loaded .env from {ENV_PATH}")
    except Exception as e:
        logging.warning(f"Failed to load .env: {e}")
    finally:
        _ENV_LOADED = True


def get_env(name: str, default: str = None) -> Optional[str]:
    """Get environment variable"""
    val = os.environ.get(name)
    return val if val else default


def get_timeout(name: str, default: int) -> int:
    """Get timeout value from environment or use default"""
    try:
        return int(get_env(name, str(default)))
    except (ValueError, TypeError):
        return default


def build_driver(headless: bool = True) -> webdriver.Chrome:
    """Create Chrome WebDriver instance"""
    opts = Options()
    
    debug = get_env("TMS_DEBUG", "").lower() in ("1", "true", "yes")
    if headless and not debug:
        opts.add_argument("--headless=new")
    
    opts.add_argument("--disable-gpu")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--window-size=1920,1080")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    
    # Set download directory to Downloads folder
    download_dir = str(Path.home() / "Downloads")
    prefs = {
        "credentials_enable_service": False,
        "profile.password_manager_enabled": False,
        "download.default_directory": download_dir,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": False
    }
    opts.add_experimental_option("prefs", prefs)
    
    chrome_bin = get_env("TMS_CHROME")
    if chrome_bin:
        opts.binary_location = chrome_bin
    
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=opts)
    
    # Configurable timeouts via environment variables
    page_timeout = get_timeout('TMS_PAGE_TIMEOUT', 180)
    script_timeout = get_timeout('TMS_SCRIPT_TIMEOUT', 180)
    implicit_wait = get_timeout('TMS_IMPLICIT_WAIT', 10)
    
    driver.set_page_load_timeout(page_timeout)
    driver.set_script_timeout(script_timeout)
    driver.implicitly_wait(implicit_wait)
    
    if debug:
        logging.info("DEBUG MODE: Browser visible")
    
    return driver


def login_tms(driver: webdriver.Chrome, username: str, password: str) -> bool:
    """
    Login to TMS system
    Returns True if successful, False otherwise
    """
    start_time = time.time()
    try:
        url = "https://home.eccang.com/#/user/home"
        logging.info(f"[LOGIN] Navigating to {url}...")
        driver.get(url)
        logging.info(f"[LOGIN] Page loaded, checking login status...")
        time.sleep(3)
        
        wait = WebDriverWait(driver, 30)
        
        # Check if already logged in by looking for "进入系统" button
        try:
            time.sleep(2)  # Give page time to load
            # Find the first enabled "进入系统" button (易面单-高级版-正式)
            enter_btn = wait.until(EC.element_to_be_clickable(
                (By.XPATH, "//div[contains(@class, 'system-item-content__btn')]/button[not(@disabled)]//span[contains(text(), '进入系统')]")))
            logging.info("[LOGIN] Already logged in, clicking '进入系统' for 易面单...")
            
            # Store current window handle
            original_window = driver.current_window_handle
            
            driver.execute_script("arguments[0].scrollIntoView(true);", enter_btn)
            time.sleep(1)
            driver.execute_script("arguments[0].click();", enter_btn)
            
            # Wait for new window/tab to open
            logging.info("[LOGIN] Waiting for TMS tab to open...")
            time.sleep(5)
            
            # Switch to the new window
            for window_handle in driver.window_handles:
                if window_handle != original_window:
                    driver.switch_to.window(window_handle)
                    elapsed = time.time() - start_time
                    logging.info(f"[LOGIN] ✓ Switched to TMS tab (took {elapsed:.1f}s)")
                    break
            
            time.sleep(3)
            return True
        except TimeoutException:
            logging.info("[LOGIN] Not logged in, proceeding with login...")
        
        # Find and fill login form
        logging.info("[LOGIN] Looking for login form...")
        username_input = wait.until(EC.presence_of_element_located((By.ID, "userName")))
        password_input = wait.until(EC.presence_of_element_located((By.ID, "password")))
        
        username_input.clear()
        username_input.send_keys(username)
        password_input.clear()
        password_input.send_keys(password)
        
        logging.info("Submitting login form")
        # Find and click login button
        login_btn = wait.until(EC.element_to_be_clickable(
            (By.CSS_SELECTOR, "a.ec-login-submit")))
        login_btn.click()
        
        time.sleep(5)
        
        # Click "进入系统" button after login
        logging.info("Clicking '进入系统' button for 易面单")
        # Wait for page to fully load
        time.sleep(2)
        
        # Store current window handle
        original_window = driver.current_window_handle
        
        # Find the first enabled "进入系统" button (not disabled, for 易面单-高级版-正式)
        enter_btn = wait.until(EC.element_to_be_clickable(
            (By.XPATH, "//div[contains(@class, 'system-item-content__btn')]/button[not(@disabled)]//span[contains(text(), '进入系统')]")))
        
        # Scroll to button to make sure it's visible
        driver.execute_script("arguments[0].scrollIntoView(true);", enter_btn)
        time.sleep(1)
        driver.execute_script("arguments[0].click();", enter_btn)
        
        # Wait for new window/tab to open
        logging.info("Waiting for new tab to open...")
        time.sleep(5)
        
        # Switch to the new window
        for window_handle in driver.window_handles:
            if window_handle != original_window:
                driver.switch_to.window(window_handle)
                logging.info("Switched to TMS tab")
                break
        
        time.sleep(3)
        
        time.sleep(5)
        elapsed = time.time() - start_time
        logging.info(f"[LOGIN] ✓ Login successful (took {elapsed:.1f}s)")
        return True
        
    except Exception as e:
        logging.error(f"Login failed: {e}")
        try:
            screenshot_path = Path(__file__).parent / f"tms_login_error_{int(time.time())}.png"
            driver.save_screenshot(str(screenshot_path))
            logging.error(f"Screenshot saved: {screenshot_path}")
        except:
            pass
        return False


def navigate_to_order_list(driver: webdriver.Chrome) -> bool:
    """
    Navigate to order list page
    Returns True if successful
    """
    start_time = time.time()
    try:
        wait = WebDriverWait(driver, 30)
        
        logging.info("[NAVIGATE] Waiting for dashboard to load...")
        time.sleep(3)
        
        # Click "订单列表" directly (menu may already be open)
        logging.info("[NAVIGATE] Looking for '订单列表' menu item...")
        try:
            # Try clicking the link directly first
            logging.info("[NAVIGATE] Attempting to find order list link directly...")
            order_list = wait.until(EC.element_to_be_clickable(
                (By.XPATH, "//a[@href='/orderManage/orderList']//span[text()='订单列表']")))
            logging.info("[NAVIGATE] Found order list link, clicking...")
            driver.execute_script("arguments[0].click();", order_list)
            logging.info("[NAVIGATE] ✓ Clicked '订单列表' directly")
        except:
            # Menu is collapsed, need to click "订单管理" to expand it first
            logging.info("[NAVIGATE] Order list not visible, expanding menu...")
            order_mgmt = wait.until(EC.element_to_be_clickable(
                (By.XPATH, "//li[contains(@class, 'ant-menu-submenu')]//span[contains(text(), '订单管理')]")))
            logging.info("[NAVIGATE] Found order management menu, clicking to expand...")
            driver.execute_script("arguments[0].click();", order_mgmt)
            time.sleep(2)
            
            # Now click "订单列表"
            logging.info("[NAVIGATE] Now looking for '订单列表' in expanded menu...")
            order_list = wait.until(EC.element_to_be_clickable(
                (By.XPATH, "//a[@href='/orderManage/orderList']//span[text()='订单列表']")))
            logging.info("[NAVIGATE] Found order list, clicking...")
            driver.execute_script("arguments[0].click();", order_list)
        
        logging.info("[NAVIGATE] Waiting for order list page to load...")
        time.sleep(5)
        elapsed = time.time() - start_time
        logging.info(f"[NAVIGATE] ✓ Order list page loaded (took {elapsed:.1f}s)")
        return True
        
    except Exception as e:
        logging.error(f"Navigation failed: {e}")
        try:
            screenshot_path = Path(__file__).parent / f"tms_nav_error_{int(time.time())}.png"
            driver.save_screenshot(str(screenshot_path))
            logging.error(f"Screenshot saved: {screenshot_path}")
        except:
            pass
        return False


def parse_tms_excel(file_path) -> List[Dict]:
    """
    Parse TMS exported Excel file
    Returns list of dictionaries with tracking data (includes ALL columns)
    """
    # Convert to Path if string
    if isinstance(file_path, str):
        file_path = Path(file_path)
    
    results = []
    
    try:
        logging.info(f"[PARSE] Loading Excel workbook: {file_path}")
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        logging.info(f"[PARSE] ✓ Workbook loaded, reading headers...")
        
        # Find header row (first row)
        raw_headers = []
        for cell in ws[1]:
            raw_headers.append(cell.value)
        
        logging.info(f"[PARSE] Found {len(raw_headers)} columns in Excel file")
        
        # Disambiguate duplicate column headers.
        # TMS exports have two address blocks with identical column names:
        #   1st block = recipient (收件人): 姓名, 公司名称, ..., 邮编, ...
        #   2nd block = sender   (发件人): 姓名, 公司名称, ..., 邮编, ...
        # Also 计费重 appears twice (order-level and box-level).
        # We prefix duplicates so both are preserved in raw_data.
        ADDRESS_FIELDS = {'姓名', '公司名称', '国家简码', '省、州', '城市', '区',
                          '街道地址1', '街道地址2', '街道地址3', '邮编', '电话', '邮箱', '门牌号'}
        headers = []
        seen_counts = {}  # header -> how many times seen so far
        for h in raw_headers:
            if h is None:
                headers.append(None)
                continue
            seen_counts[h] = seen_counts.get(h, 0) + 1
            count = seen_counts[h]
            if count == 1:
                # First occurrence: if it's an address field, prefix with 收件人_
                # (only when there IS a duplicate later — we'll fix in a 2nd pass)
                headers.append(h)
            elif count == 2:
                if h in ADDRESS_FIELDS:
                    headers.append(f'发件人_{h}')  # 2nd address block = sender
                elif h == '计费重':
                    headers.append('箱计费重')  # 2nd 计费重 is per-box
                else:
                    headers.append(f'{h}_2')
            else:
                headers.append(f'{h}_{count}')
        
        # 2nd pass: rename the FIRST occurrence of address fields to 收件人_ prefix
        # only if a 2nd occurrence was seen (i.e. the export has duplicate columns)
        for h in ADDRESS_FIELDS:
            if seen_counts.get(h, 0) >= 2:
                for i, hdr in enumerate(headers):
                    if hdr == h:
                        headers[i] = f'收件人_{h}'
                        break  # only rename the first one
        
        logging.info(f"[PARSE] Headers disambiguated. Duplicates handled: {[h for h,c in seen_counts.items() if c > 1]}")
        
        # Map column names to indices
        col_map = {header: idx for idx, header in enumerate(headers) if header}
        
        # Required columns for quick access
        tracking_col = col_map.get('子单号')
        customer_col = col_map.get('商户名称')
        api_cost_col = col_map.get('应付金额')
        charged_col = col_map.get('扣款金额')
        master_tracking_col = col_map.get('服务商单号')  # Master tracking number
        
        # Find the order-level 计费重 (may have been left as-is or renamed)
        weight_col = col_map.get('计费重')
        if weight_col is None:
            # Fallback: find first occurrence by scanning
            for idx, header in enumerate(headers):
                if header and '计费重' in header:
                    weight_col = idx
                    break
        
        logging.info(f"[PARSE] Column mapping: 子单号={tracking_col}, 服务商单号={master_tracking_col}, 商户名称={customer_col}, 应付金额={api_cost_col}, 扣款金额={charged_col}, 计费重={weight_col}")
        
        if tracking_col is None:
            logging.error(f"[PARSE] ✗ Required column '子单号' not found in Excel")
            logging.error(f"[PARSE] Available columns: {list(col_map.keys())[:10]}...") 
            raise Exception("Required column '子单号' not found in Excel")
        
        # Parse data rows
        logging.info(f"[PARSE] Parsing data rows from Excel (storing all {len(headers)} columns)...")
        row_count = 0
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[tracking_col]:
                continue
            
            row_count += 1
            
            tracking_number = str(row[tracking_col]).strip()
            if not tracking_number:
                continue
            
            # Create record with frequently used fields
            record = {
                'tracking_number': tracking_number,
                'master_tracking_number': str(row[master_tracking_col]) if master_tracking_col is not None and row[master_tracking_col] else tracking_number,
                'customer_name': str(row[customer_col]) if customer_col is not None and row[customer_col] else '',
                'api_cost': float(row[api_cost_col]) if api_cost_col is not None and row[api_cost_col] else 0.0,
                'charged_amount': float(row[charged_col]) if charged_col is not None and row[charged_col] else 0.0,
                'weight_kg': float(row[weight_col]) if weight_col is not None and row[weight_col] else 0.0
            }
            
            # Store ALL columns for future use
            for idx, header in enumerate(headers):
                if header and header not in record:  # Don't overwrite main fields
                    value = row[idx]
                    # Convert to appropriate type
                    if value is None:
                        record[header] = None
                    elif isinstance(value, (int, float, bool)):
                        record[header] = value
                    else:
                        record[header] = str(value)
            
            results.append(record)
        
        wb.close()
        logging.info(f"[PARSE] Parsed {len(results)} records from {file_path.name}")
        logging.info(f"[PARSE] Each record contains all {len(headers)} columns from TMS export")
        
        return results
        
    except Exception as e:
        logging.error(f"Failed to parse Excel file: {e}")
        return []


def test_connection(username: str = None, password: str = None) -> bool:
    """
    Test TMS connection
    Returns True if connection successful
    """
    load_env_file()
    
    username = username or get_env("TMS_USER")
    password = password or get_env("TMS_PASS")
    
    if not username or not password:
        logging.error("TMS credentials not configured")
        return False
    
    driver = None
    try:
        logging.info("Testing TMS connection")
        driver = build_driver(headless=True)
        
        if not login_tms(driver, username, password):
            return False
        
        if not navigate_to_order_list(driver):
            return False
        
        logging.info("✓ TMS connection test successful")
        return True
        
    except Exception as e:
        logging.error(f"TMS connection test failed: {e}")
        return False
    finally:
        if driver:
            try:
                driver.quit()
            except:
                pass



